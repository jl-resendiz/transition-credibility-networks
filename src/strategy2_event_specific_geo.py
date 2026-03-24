"""Event-specific geographic weight matrix test.

The baseline w_geo uses MW-weighted centroids, which dilutes geographic
signal for multinationals. This script rebuilds w_geo for each retirement
event: distance is measured from the RETIRING PLANT to each firm's NEAREST
plant (not centroid).

Three specifications:
  A (baseline):       CAR = b1*w_fuel + b2*w_geo_CENTROID + b3*w_reg + b4*same_sector
  B (event-specific): CAR = b1*w_fuel + b2*w_geo_EVENT   + b3*w_reg + b4*same_sector
  C (both):           CAR = b1*w_fuel + b2*w_geo_CENTROID + b3*w_geo_EVENT + b4*w_reg + b5*same_sector

Reports pooled event-clustered AND Fama-MacBeth + Newey-West for all three.
Output: results/metrics/strategy2_event_specific_geo.md
"""
import csv
import os
import sys
import math
import random
import hashlib
import re
import openpyxl
from collections import defaultdict

from _paths import derived_path, raw_path, results_path


def _print(msg=''):
    print(msg)
    sys.stdout.flush()


# ── Haversine ──────────────────────────────────────────────────────

def haversine(lat1, lon1, lat2, lon2):
    """Great-circle distance in km."""
    R = 6371
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (math.sin(dlat / 2) ** 2
         + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2))
         * math.sin(dlon / 2) ** 2)
    return R * 2 * math.asin(math.sqrt(a))


# ── Matrix utilities ──────────────────────────────────────────────

def invert_matrix(mat):
    n = len(mat)
    aug = [row[:] + [1.0 if i == j else 0.0 for j in range(n)]
           for i, row in enumerate(mat)]
    for col in range(n):
        max_row = max(range(col, n), key=lambda r: abs(aug[r][col]))
        if abs(aug[max_row][col]) < 1e-20:
            return None
        aug[col], aug[max_row] = aug[max_row], aug[col]
        pivot = aug[col][col]
        for j in range(2 * n):
            aug[col][j] /= pivot
        for row in range(n):
            if row != col:
                factor = aug[row][col]
                for j in range(2 * n):
                    aug[row][j] -= factor * aug[col][j]
    return [row[n:] for row in aug]


def mat_mul(a, b):
    rows_a, cols_b, mid = len(a), len(b[0]), len(b)
    out = [[0.0] * cols_b for _ in range(rows_a)]
    for i in range(rows_a):
        for k in range(mid):
            if a[i][k] == 0:
                continue
            for j in range(cols_b):
                out[i][j] += a[i][k] * b[k][j]
    return out


# ── OLS ───────────────────────────────────────────────────────────

def ols_simple(y, X_mat):
    n = len(y)
    k = len(X_mat[0])
    if n <= k:
        return None
    y_mean = sum(y) / n
    ss_tot = sum((yi - y_mean) ** 2 for yi in y)
    if ss_tot < 1e-15:
        return None
    XtX = [[sum(X_mat[i][a] * X_mat[i][b] for i in range(n))
            for b in range(k)] for a in range(k)]
    Xty = [sum(X_mat[i][a] * y[i] for i in range(n)) for a in range(k)]
    inv_XtX = invert_matrix(XtX)
    if inv_XtX is None:
        return None
    beta = [sum(inv_XtX[a][b] * Xty[b] for b in range(k)) for a in range(k)]
    y_hat = [sum(X_mat[i][a] * beta[a] for a in range(k)) for i in range(n)]
    resid = [y[i] - y_hat[i] for i in range(n)]
    ss_res = sum(r ** 2 for r in resid)
    r2 = 1 - ss_res / ss_tot
    return {'beta': beta, 'resid': resid, 'r2': r2, 'n': n,
            'inv_XtX': inv_XtX, 'X': X_mat, 'y': y}


def ols_with_clustered_se(y, X_mat, cluster_ids):
    """OLS with one-way clustered standard errors."""
    result = ols_simple(y, X_mat)
    if result is None:
        return None
    n = result['n']
    k = len(result['beta'])
    resid = result['resid']
    inv_XtX = result['inv_XtX']

    # Cluster-robust meat
    clusters = defaultdict(list)
    for i, cid in enumerate(cluster_ids):
        clusters[cid].append(i)
    G = len(clusters)

    meat = [[0.0] * k for _ in range(k)]
    for cid, indices in clusters.items():
        score = [sum(X_mat[i][a] * resid[i] for i in indices) for a in range(k)]
        for a in range(k):
            for b in range(k):
                meat[a][b] += score[a] * score[b]

    # Small-sample correction
    correction = (G / (G - 1.0)) * ((n - 1.0) / (n - k))

    # V = inv(XtX) * meat * inv(XtX) * correction
    tmp = mat_mul(inv_XtX, meat)
    V = mat_mul(tmp, inv_XtX)
    se = [math.sqrt(max(V[a][a] * correction, 0)) for a in range(k)]
    t_stats = [result['beta'][a] / se[a] if se[a] > 1e-15 else 0.0
               for a in range(k)]
    result['se'] = se
    result['t'] = t_stats
    result['n_clusters'] = G
    return result


# ── Normal CDF / p-value ──────────────────────────────────────────

def _normal_cdf(x):
    if x < -8:
        return 0.0
    if x > 8:
        return 1.0
    ax = abs(x)
    b0 = 0.2316419
    b1 = 0.319381530
    b2 = -0.356563782
    b3 = 1.781477937
    b4 = -1.821255978
    b5 = 1.330274429
    t_val = 1.0 / (1.0 + b0 * ax)
    phi = (1.0 / math.sqrt(2.0 * math.pi)) * math.exp(-0.5 * ax * ax)
    cdf = 1.0 - phi * (b1 * t_val + b2 * t_val ** 2 + b3 * t_val ** 3
                        + b4 * t_val ** 4 + b5 * t_val ** 5)
    return 1.0 - cdf if x < 0 else cdf


def p_from_t(t_stat):
    return 2.0 * (1.0 - _normal_cdf(abs(t_stat)))


# ── Newey-West HAC SE ─────────────────────────────────────────────

def newey_west_se(series, max_lag=None):
    T = len(series)
    if T < 3:
        return float('inf')
    mean = sum(series) / T
    demean = [x - mean for x in series]
    if max_lag is None:
        max_lag = max(1, int(4 * (T / 100) ** (2 / 9)))
    max_lag = min(max_lag, T - 1)
    gamma_0 = sum(d * d for d in demean) / T
    nw_var = gamma_0
    for lag in range(1, max_lag + 1):
        weight = 1.0 - lag / (max_lag + 1.0)
        gamma_lag = sum(demean[t] * demean[t - lag] for t in range(lag, T)) / T
        nw_var += 2.0 * weight * gamma_lag
    var_mean = nw_var / T
    if var_mean < 0:
        var_mean = gamma_0 / T
    return math.sqrt(var_mean)


# ══════════════════════════════════════════════════════════════════
# LOAD DATA
# ══════════════════════════════════════════════════════════════════

_print('=' * 70)
_print('EVENT-SPECIFIC GEOGRAPHIC WEIGHT MATRIX')
_print('=' * 70)

# ── 1. Load GEM->Compustat parent mapping ─────────────────────────

def parse_parents(field):
    if not field or str(field).strip() == '':
        return []
    parts = str(field).split(';')
    results = []
    for p in parts:
        p = p.strip()
        match = re.match(r'^(.+?)\s*\[(\d+\.?\d*)%\]$', p)
        if match:
            results.append((match.group(1).strip(), float(match.group(2))))
        elif p:
            results.append((p.strip(), None))
    return results


parent_to_gvkeys = defaultdict(set)
with open(derived_path('mappings', 'gem_compustat_matches.csv'), 'r',
          encoding='utf-8') as f:
    for row in csv.DictReader(f):
        parent_to_gvkeys[row['gem_parent']].add(row['gvkey'])

_print(f'Matched parents: {len(parent_to_gvkeys)}')

# ── 2. Read plant-level GPS from GEM trackers ────────────────────

_print('Loading plant-level GPS from GEM trackers...')
gvkey_plants = defaultdict(list)  # gvkey -> [(lat, lon, mw), ...]

trackers = [
    ('Global-Coal-Plant-Tracker-January-2026.xlsx', 'Units',
     'Parent', 'Capacity (MW)', 'Latitude', 'Longitude'),
    ('Global-Oil-and-Gas-Plant-Tracker-GOGPT-January-2026.xlsx',
     'Gas & Oil Units', 'Parent(s)', 'Capacity (MW)', 'Latitude', 'Longitude'),
    ('Global-Solar-Power-Tracker-February-2026.xlsx',
     'Utility-Scale (1 MW+)', 'Owner', 'Capacity (MW)', 'Latitude', 'Longitude'),
    ('Global-Wind-Power-Tracker-February-2026.xlsx',
     'Data', 'Owner', 'Capacity (MW)', 'Latitude', 'Longitude'),
]

for fname, sheet, parent_col, cap_col, lat_col, lon_col in trackers:
    fpath = raw_path('gem', fname)
    if not os.path.exists(fpath):
        _print(f'  WARNING: {fname} not found, skipping')
        continue
    _print(f'  Reading {fname}...')
    wb = openpyxl.load_workbook(fpath, read_only=True)
    ws = wb[sheet]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    parent_idx = headers.index(parent_col)
    cap_idx = headers.index(cap_col)
    lat_idx = headers.index(lat_col)
    lon_idx = headers.index(lon_col)
    status_idx = headers.index('Status')

    for row in ws.iter_rows(min_row=2, values_only=True):
        status = str(row[status_idx]) if row[status_idx] else ''
        if status != 'operating':
            continue
        try:
            cap = float(row[cap_idx])
            lat = float(row[lat_idx])
            lon = float(row[lon_idx])
        except (ValueError, TypeError):
            continue
        parsed = parse_parents(row[parent_idx])
        for name, pct in parsed:
            if name in parent_to_gvkeys:
                share = ((pct / 100.0) if pct
                         else 1.0 / len(parsed) if len(parsed) > 1 else 1.0)
                for gvkey in parent_to_gvkeys[name]:
                    gvkey_plants[gvkey].append((lat, lon, cap * share))
    wb.close()

_print(f'Firms with plant GPS data: {len(gvkey_plants)}')

# ── 3. Compute centroids (for baseline Spec A) ───────────────────

centroids = {}
for gvkey, plants in gvkey_plants.items():
    total_mw = sum(mw for _, _, mw in plants)
    if total_mw <= 0:
        continue
    wlat = sum(lat * mw for lat, lon, mw in plants) / total_mw
    wlon = sum(lon * mw for lat, lon, mw in plants) / total_mw
    centroids[gvkey] = (wlat, wlon)

_print(f'Firms with valid centroids: {len(centroids)}')

# ── 4. Load monthly returns ──────────────────────────────────────

_print('Loading monthly returns...')
monthly_ret = defaultdict(dict)
with open(derived_path('returns', 'monthly_returns.csv'), 'r',
          encoding='utf-8') as f:
    for row in csv.DictReader(f):
        gk = row['gvkey']
        ym = row['datadate'][:7]
        try:
            monthly_ret[gk][ym] = float(row['ret_monthly'])
        except ValueError:
            pass
_print(f'  Monthly: {len(monthly_ret)} firms')


def load_ff_factors_monthly(path):
    if not os.path.exists(path):
        return None
    vwretd = {}
    with open(path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('This file') or line.startswith('The '):
                continue
            if line.startswith(','):
                continue
            parts = [p.strip() for p in line.split(',')]
            if len(parts) < 5:
                continue
            date = parts[0]
            if not date.isdigit() or len(date) != 6:
                continue
            try:
                mktrf_val = float(parts[1])
                rf_val = float(parts[4])
            except ValueError:
                continue
            vw = (mktrf_val + rf_val) / 100.0
            vwretd[f'{date[:4]}-{date[4:6]}'] = vw
    return vwretd


market_ret_monthly = load_ff_factors_monthly(
    raw_path('factors', 'F-F_Research_Data_Factors.csv')
)
if not market_ret_monthly:
    raise RuntimeError('Missing F-F monthly factors.')
_print(f'  Market months: {len(market_ret_monthly)}')

# ── 5. Load weight matrices (fuel, regulatory, centroid-geo) ─────

_print('Loading weight matrices...')
W_geo_centroid = defaultdict(dict)
geo_path = derived_path('networks', 'weight_matrix_W_geo.csv')
if os.path.exists(geo_path):
    with open(geo_path, 'r', encoding='utf-8') as f:
        for row in csv.DictReader(f):
            W_geo_centroid[row['gvkey_i']][row['gvkey_j']] = float(row['w_ij'])
_print(f'  W_geo_centroid: {len(W_geo_centroid)} firms')

W_fuel = defaultdict(dict)
fuel_path = derived_path('networks', 'weight_matrix_W_fuel.csv')
if os.path.exists(fuel_path):
    with open(fuel_path, 'r', encoding='utf-8') as f:
        for row in csv.DictReader(f):
            W_fuel[row['gvkey_i']][row['gvkey_j']] = float(row['w_ij'])
_print(f'  W_fuel: {len(W_fuel)} firms')

W_reg = defaultdict(dict)
reg_path = derived_path('networks', 'weight_matrix_W_regulatory.csv')
if os.path.exists(reg_path):
    with open(reg_path, 'r', encoding='utf-8') as f:
        for row in csv.DictReader(f):
            wval = row.get('w_ij')
            if wval in (None, ''):
                wval = row.get('w_reg')
            try:
                W_reg[row['gvkey_i']][row['gvkey_j']] = float(wval)
            except (ValueError, TypeError):
                continue
_print(f'  W_reg: {len(W_reg)} firms')

# ── 6. Load fundamentals (for SIC codes) ─────────────────────────

fundamentals = {}
with open(derived_path('fundamentals', 'firm_fundamentals.csv'), 'r',
          encoding='utf-8') as f:
    for row in csv.DictReader(f):
        gk = row['gvkey']
        fy = row['fyear']
        if gk not in fundamentals or fy > fundamentals[gk]['fyear']:
            fundamentals[gk] = row


def get_sic4(gvkey):
    f = fundamentals.get(gvkey)
    if f and f.get('sic'):
        return f['sic'][:4]
    return None


# ── 7. Load retirement events ────────────────────────────────────

_print('Loading events...')
all_events = []
with open(derived_path('events', 'coal_retirement_events.csv'), 'r',
          encoding='utf-8') as f:
    for row in csv.DictReader(f):
        if not row.get('matched_gvkeys'):
            continue
        if row.get('is_first_mover') != 'True':
            continue
        ann_date = row.get('announcement_date', '').strip()
        ret_date = row.get('event_date', '').strip()
        effective_date = ann_date if ann_date else ret_date
        event_year = None
        if (effective_date and len(effective_date) >= 4
                and effective_date[:4].isdigit()):
            event_year = int(effective_date[:4])
        else:
            event_year = int(row['ret_year']) if row.get('ret_year') else None

        # Retiring plant GPS
        try:
            ev_lat = float(row['lat'])
            ev_lon = float(row['lon'])
        except (ValueError, TypeError, KeyError):
            ev_lat, ev_lon = None, None

        all_events.append({
            'gem_id': row.get('gem_id', ''),
            'plant': row['plant_name'],
            'year': event_year,
            'event_date': effective_date,
            'gvkeys': row['matched_gvkeys'].split(';'),
            'lat': ev_lat,
            'lon': ev_lon,
        })

_print(f'  First-mover events: {len(all_events)}')
n_with_gps = sum(1 for e in all_events if e['lat'] is not None)
_print(f'  Events with GPS: {n_with_gps}')


# ══════════════════════════════════════════════════════════════════
# BUILD EVENT-SPECIFIC W_GEO
# ══════════════════════════════════════════════════════════════════

POST_MONTHS = 3
PRE_MONTHS = 24
DECAY_KM = 1000 / math.log(2)  # half-life 1000 km
MIN_OBS_PER_EVENT = 20


def compute_monthly_car(gvkey, event_month, post=3):
    if gvkey not in monthly_ret:
        return None
    months = sorted(monthly_ret[gvkey].keys())
    event_idx = None
    for i, m in enumerate(months):
        if m >= event_month:
            event_idx = i
            break
    if event_idx is None:
        return None
    ar_list = []
    for i in range(max(0, event_idx - PRE_MONTHS), event_idx):
        m = months[i]
        if m in monthly_ret[gvkey] and m in market_ret_monthly:
            ar_list.append(monthly_ret[gvkey][m] - market_ret_monthly[m])
    if len(ar_list) < 12:
        return None
    pre_mean_ar = sum(ar_list) / len(ar_list)
    car = 0.0
    for offset in range(-1, post + 1):
        idx = event_idx + offset
        if 0 <= idx < len(months) and months[idx] in monthly_ret[gvkey]:
            m = months[idx]
            r_it = monthly_ret[gvkey][m]
            if m in market_ret_monthly:
                ar = r_it - market_ret_monthly[m]
                car += ar - pre_mean_ar
    return car


def nearest_plant_distance(ev_lat, ev_lon, gvkey):
    """Distance from event plant to firm's nearest plant."""
    plants = gvkey_plants.get(gvkey)
    if not plants:
        return None
    min_d = float('inf')
    for lat, lon, mw in plants:
        d = haversine(ev_lat, ev_lon, lat, lon)
        if d < min_d:
            min_d = d
    return min_d if min_d < float('inf') else None


def event_specific_geo_weight(ev_lat, ev_lon, firm_gvkeys):
    """Build row-normalized w_geo_event for all firms, given event GPS.

    Returns dict gvkey -> weight.
    """
    raw = {}
    for gk in firm_gvkeys:
        d = nearest_plant_distance(ev_lat, ev_lon, gk)
        if d is not None and d > 0:
            raw[gk] = math.exp(-d / DECAY_KM) / d
    total = sum(raw.values())
    if total <= 0:
        return {}
    return {gk: w / total for gk, w in raw.items()}


# ══════════════════════════════════════════════════════════════════
# BUILD PER-EVENT DATASETS
# ══════════════════════════════════════════════════════════════════

_print('\nBuilding per-event datasets with event-specific w_geo...')

# Collect all gvkeys that appear in any weight matrix
all_gvkeys = set(W_fuel.keys()) | set(W_geo_centroid.keys()) | set(W_reg.keys())
all_gvkeys |= set(fundamentals.keys())
# Keep only those with return data
all_gvkeys = {gk for gk in all_gvkeys if gk in monthly_ret}
_print(f'  Candidate firms: {len(all_gvkeys)}')

SPEC_A_VARS = ['w_fuel', 'w_geo_centroid', 'w_reg', 'same_sector']
SPEC_B_VARS = ['w_fuel', 'w_geo_event', 'w_reg', 'same_sector']
SPEC_C_VARS = ['w_fuel', 'w_geo_centroid', 'w_geo_event', 'w_reg', 'same_sector']

event_datasets = {}  # event_id -> list of obs dicts

for event_id, event in enumerate(all_events):
    ev_lat = event['lat']
    ev_lon = event['lon']
    if ev_lat is None or ev_lon is None:
        continue

    event_gvkeys = set(event['gvkeys'])
    year = event['year']
    event_date = event.get('event_date', '')
    if event_date and len(event_date) >= 7:
        event_month = event_date[:7]
    else:
        event_month = f'{year}-07' if year else None
    if not event_month:
        continue

    fm_sic4 = None
    for gk in event_gvkeys:
        fm_sic4 = get_sic4(gk)
        if fm_sic4:
            break

    # Compute event-specific geo weights for ALL firms
    w_geo_ev = event_specific_geo_weight(ev_lat, ev_lon,
                                         all_gvkeys - event_gvkeys)

    obs = []
    for fm_gk in event_gvkeys:
        if fm_gk not in W_geo_centroid:
            continue
        neighbors_centroid = W_geo_centroid[fm_gk]
        neighbor_gks = set(neighbors_centroid.keys()) - event_gvkeys

        # Also include firms that have event-specific geo > 0
        neighbor_gks |= set(w_geo_ev.keys()) - event_gvkeys

        non_connected = [gk for gk in all_gvkeys
                         if gk not in event_gvkeys
                         and gk not in neighbor_gks]
        stable_seed = int(hashlib.md5(
            str(fm_gk).encode('utf-8')).hexdigest()[:8], 16)
        random.seed(stable_seed)
        n_ctrl = min(len(non_connected),
                     max(5 * len(neighbor_gks), 20))
        ctrl_sample = (random.sample(non_connected, n_ctrl)
                       if len(non_connected) > n_ctrl
                       else non_connected)
        candidate_firms = list(neighbor_gks) + ctrl_sample

        for gk in candidate_firms:
            w_fuel_val = W_fuel.get(fm_gk, {}).get(gk, 0.0)
            w_geo_c = neighbors_centroid.get(gk, 0.0)
            w_geo_e = w_geo_ev.get(gk, 0.0)
            w_reg_val = W_reg.get(fm_gk, {}).get(gk, 0.0)
            j_sic4 = get_sic4(gk)
            same_sector = 1.0 if (fm_sic4 and j_sic4
                                  and fm_sic4 == j_sic4) else 0.0
            car = compute_monthly_car(gk, event_month, post=POST_MONTHS)
            if car is None:
                continue
            obs.append({
                'car': car,
                'w_fuel': w_fuel_val,
                'w_geo_centroid': w_geo_c,
                'w_geo_event': w_geo_e,
                'w_reg': w_reg_val,
                'same_sector': same_sector,
                'gvkey': gk,
                'event_id': event_id,
            })

    if len(obs) >= MIN_OBS_PER_EVENT:
        event_datasets[event_id] = obs

n_valid = len(event_datasets)
_print(f'  Valid events (>= {MIN_OBS_PER_EVENT} obs): {n_valid}')
_print(f'  Total obs: {sum(len(v) for v in event_datasets.values())}')

# Diagnostics: correlation between centroid and event-specific geo
all_obs = [o for ds in event_datasets.values() for o in ds]
n_both_pos = sum(1 for o in all_obs
                 if o['w_geo_centroid'] > 0 and o['w_geo_event'] > 0)
n_any = len(all_obs)
mean_centroid = sum(o['w_geo_centroid'] for o in all_obs) / n_any if n_any else 0
mean_event = sum(o['w_geo_event'] for o in all_obs) / n_any if n_any else 0
_print(f'\n  Diagnostics:')
_print(f'    Mean w_geo_centroid: {mean_centroid:.6f}')
_print(f'    Mean w_geo_event:    {mean_event:.6f}')
_print(f'    Both > 0:            {n_both_pos}/{n_any} '
       f'({100*n_both_pos/n_any:.1f}%)' if n_any > 0 else '')

# Correlation
if n_any > 1:
    cx = [o['w_geo_centroid'] for o in all_obs]
    cy = [o['w_geo_event'] for o in all_obs]
    mx, my = mean_centroid, mean_event
    cov_xy = sum((cx[i] - mx) * (cy[i] - my) for i in range(n_any)) / n_any
    sd_x = math.sqrt(sum((x - mx) ** 2 for x in cx) / n_any)
    sd_y = math.sqrt(sum((y - my) ** 2 for y in cy) / n_any)
    corr = cov_xy / (sd_x * sd_y) if sd_x > 0 and sd_y > 0 else 0
    _print(f'    Correlation:         {corr:.4f}')


# ══════════════════════════════════════════════════════════════════
# POOLED EVENT-CLUSTERED REGRESSIONS
# ══════════════════════════════════════════════════════════════════

_print('\n' + '=' * 70)
_print('POOLED REGRESSIONS (event-clustered SEs)')
_print('=' * 70)

specs = [
    ('A: Centroid w_geo', SPEC_A_VARS),
    ('B: Event-specific w_geo', SPEC_B_VARS),
    ('C: Both geo measures', SPEC_C_VARS),
]

pooled_results = {}
for spec_name, var_names in specs:
    y = [o['car'] for ds in event_datasets.values() for o in ds]
    X = [[1.0] + [o[v] for v in var_names]
         for ds in event_datasets.values() for o in ds]
    cluster_ids = [o['event_id']
                   for ds in event_datasets.values() for o in ds]

    result = ols_with_clustered_se(y, X, cluster_ids)
    if result is None:
        _print(f'\n  {spec_name}: OLS failed')
        continue

    names = ['intercept'] + var_names
    _print(f'\n  {spec_name} (N={result["n"]}, clusters={result["n_clusters"]}, '
           f'R2={result["r2"]:.4f})')
    _print(f'  {"Variable":<18} {"beta":>12} {"SE":>10} {"t":>8}')
    _print('  ' + '-' * 50)
    spec_detail = {}
    for i, name in enumerate(names):
        b = result['beta'][i]
        se = result['se'][i]
        t = result['t'][i]
        p = p_from_t(t)
        stars = '***' if p < 0.01 else '**' if p < 0.05 else '*' if p < 0.10 else ''
        _print(f'  {name:<18} {b:+12.6f} {se:10.6f} {t:8.3f}{stars}')
        spec_detail[name] = {'beta': b, 'se': se, 't': t, 'p': p}
    pooled_results[spec_name] = {
        'detail': spec_detail,
        'n': result['n'],
        'n_clusters': result['n_clusters'],
        'r2': result['r2'],
    }


# ══════════════════════════════════════════════════════════════════
# FAMA-MACBETH + NEWEY-WEST
# ══════════════════════════════════════════════════════════════════

_print('\n' + '=' * 70)
_print('FAMA-MACBETH (1973) + NEWEY-WEST SEs')
_print('=' * 70)

fm_all = {}

for spec_name, var_names in specs:
    event_betas = defaultdict(list)
    event_r2s = []
    event_ns = []
    event_ids_used = []

    for event_id in sorted(event_datasets.keys()):
        obs = event_datasets[event_id]

        # Check same_sector variation
        use_vars = list(var_names)
        ss_vals = set(o['same_sector'] for o in obs)
        if len(ss_vals) <= 1 and 'same_sector' in use_vars:
            use_vars = [v for v in use_vars if v != 'same_sector']

        y = [o['car'] for o in obs]
        X = [[1.0] + [o[v] for v in use_vars] for o in obs]

        result = ols_simple(y, X)
        if result is None:
            continue

        names_ev = ['intercept'] + use_vars
        for i, name in enumerate(names_ev):
            event_betas[name].append(result['beta'][i])
        # Pad missing
        for v in var_names:
            if v not in use_vars:
                event_betas[v].append(float('nan'))

        event_r2s.append(result['r2'])
        event_ns.append(result['n'])
        event_ids_used.append(event_id)

    T_fm = len(event_ids_used)
    if T_fm < 3:
        _print(f'\n  {spec_name}: only {T_fm} valid events, skipping')
        continue

    _print(f'\n  {spec_name}')
    _print(f'    Events: {T_fm}, Avg N/event: {sum(event_ns)/T_fm:.1f}, '
           f'Avg R2: {sum(event_r2s)/T_fm:.4f}')
    _print(f'    {"Variable":<18} {"Mean beta":>12} {"NW SE":>10} {"t":>8} {"p":>8}')
    _print('    ' + '-' * 60)

    fm_spec = {}
    for v in ['intercept'] + list(var_names):
        betas = event_betas.get(v, [])
        clean = [b for b in betas if not math.isnan(b)]
        if len(clean) < 3:
            continue
        mean_b = sum(clean) / len(clean)
        nw_se = newey_west_se(clean)
        t_stat = mean_b / nw_se if nw_se > 1e-15 else 0.0
        p_val = p_from_t(t_stat)
        stars = ('***' if p_val < 0.01 else '**' if p_val < 0.05
                 else '*' if p_val < 0.10 else '')
        _print(f'    {v:<18} {mean_b:+12.6f} {nw_se:10.6f} {t_stat:8.3f} '
               f'{p_val:8.4f}{stars}')
        fm_spec[v] = {'mean': mean_b, 'se': nw_se, 't': t_stat, 'p': p_val,
                       'n_events': len(clean)}

    fm_all[spec_name] = {
        'results': fm_spec,
        'T': T_fm,
        'avg_n': sum(event_ns) / T_fm,
        'avg_r2': sum(event_r2s) / T_fm,
    }


# ══════════════════════════════════════════════════════════════════
# WRITE OUTPUT
# ══════════════════════════════════════════════════════════════════

out_path = results_path('metrics', 'strategy2_event_specific_geo.md')
os.makedirs(os.path.dirname(out_path), exist_ok=True)

lines = [
    '# Event-Specific Geographic Weight Matrix',
    '',
    'Baseline w_geo uses MW-weighted centroids, diluting geographic signal',
    'for multinationals. Event-specific w_geo measures distance from the',
    'RETIRING PLANT to each firm\'s NEAREST plant.',
    '',
    f'Decay half-life: {1000} km (scale = {DECAY_KM:.1f} km)',
    f'Events: {len(all_events)} first-mover coal retirements '
    f'({n_with_gps} with GPS)',
    f'Valid events (>= {MIN_OBS_PER_EVENT} obs): {n_valid}',
    f'Total obs: {sum(len(v) for v in event_datasets.values())}',
    '',
    '## Diagnostics',
    '',
    f'Mean w_geo_centroid: {mean_centroid:.6f}',
    f'Mean w_geo_event: {mean_event:.6f}',
]
if n_any > 1:
    lines.append(f'Correlation(centroid, event): {corr:.4f}')
lines += ['']

# Pooled results
lines += [
    '## Pooled Regressions (event-clustered SEs)',
    '',
]

for spec_name, var_names in specs:
    if spec_name not in pooled_results:
        continue
    pr = pooled_results[spec_name]
    lines += [
        f'### {spec_name}',
        '',
        f'N = {pr["n"]}, clusters = {pr["n_clusters"]}, R2 = {pr["r2"]:.4f}',
        '',
        '| Variable | beta | SE | t | p |',
        '|---|---:|---:|---:|---:|',
    ]
    for v in ['intercept'] + list(var_names):
        if v in pr['detail']:
            d = pr['detail'][v]
            stars = ('***' if d['p'] < 0.01 else '**' if d['p'] < 0.05
                     else '*' if d['p'] < 0.10 else '')
            lines.append(
                f'| {v} | {d["beta"]:+.6f} | {d["se"]:.6f} '
                f'| {d["t"]:.3f} | {d["p"]:.4f}{stars} |')
    lines.append('')

# Fama-MacBeth results
lines += [
    '## Fama-MacBeth (1973) + Newey-West SEs',
    '',
]

for spec_name, var_names in specs:
    if spec_name not in fm_all:
        continue
    fm = fm_all[spec_name]
    lines += [
        f'### {spec_name}',
        '',
        f'Events: {fm["T"]}, Avg N/event: {fm["avg_n"]:.1f}, '
        f'Avg R2: {fm["avg_r2"]:.4f}',
        '',
        '| Variable | Mean beta | NW SE | t | p |',
        '|---|---:|---:|---:|---:|',
    ]
    for v in ['intercept'] + list(var_names):
        if v in fm['results']:
            r = fm['results'][v]
            stars = ('***' if r['p'] < 0.01 else '**' if r['p'] < 0.05
                     else '*' if r['p'] < 0.10 else '')
            lines.append(
                f'| {v} | {r["mean"]:+.6f} | {r["se"]:.6f} '
                f'| {r["t"]:.3f} | {r["p"]:.4f}{stars} |')
    lines.append('')

# Summary comparison table
lines += [
    '## Summary: Centroid vs Event-Specific Geography',
    '',
    '| Spec | Method | w_geo_centroid t | w_geo_event t | w_fuel t | R2 / Avg R2 |',
    '|---|---|---:|---:|---:|---:|',
]

for spec_name, var_names in specs:
    # Pooled
    if spec_name in pooled_results:
        pr = pooled_results[spec_name]
        gc_t = pr['detail'].get('w_geo_centroid', {}).get('t', '')
        ge_t = pr['detail'].get('w_geo_event', {}).get('t', '')
        wf_t = pr['detail'].get('w_fuel', {}).get('t', '')
        gc_str = f'{gc_t:.3f}' if isinstance(gc_t, float) else ''
        ge_str = f'{ge_t:.3f}' if isinstance(ge_t, float) else ''
        wf_str = f'{wf_t:.3f}' if isinstance(wf_t, float) else ''
        lines.append(f'| {spec_name} | Pooled | {gc_str} | {ge_str} | '
                     f'{wf_str} | {pr["r2"]:.4f} |')
    # FM
    if spec_name in fm_all:
        fm = fm_all[spec_name]
        gc_t = fm['results'].get('w_geo_centroid', {}).get('t', '')
        ge_t = fm['results'].get('w_geo_event', {}).get('t', '')
        wf_t = fm['results'].get('w_fuel', {}).get('t', '')
        gc_str = f'{gc_t:.3f}' if isinstance(gc_t, float) else ''
        ge_str = f'{ge_t:.3f}' if isinstance(ge_t, float) else ''
        wf_str = f'{wf_t:.3f}' if isinstance(wf_t, float) else ''
        lines.append(f'| {spec_name} | FM+NW | {gc_str} | {ge_str} | '
                     f'{wf_str} | {fm["avg_r2"]:.4f} |')

lines += [
    '',
    '## Interpretation',
    '',
    'If event-specific w_geo (nearest-plant distance) rescues the geographic',
    'channel, we expect Spec B to show a larger and more significant coefficient',
    'on w_geo_event than Spec A shows on w_geo_centroid. In Spec C (horse race),',
    'if w_geo_event dominates w_geo_centroid, the centroid measure was indeed',
    'diluted by multinational dispersion.',
]

with open(out_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

_print(f'\nWrote: {out_path}')
_print('Done.')
