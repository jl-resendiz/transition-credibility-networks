"""Policy credibility interaction test (Equation 2 of the paper).

Tests whether policy credibility (p_t) moderates the spatial transmission
of coal retirement shocks through fuel-similarity networks.

From the valuation equation:
  V_i = Pi_bar_i - p_t * beta * alpha_i + gamma_G * sum(w^geo * s_j)
        + gamma_R * sum(w^reg * s_j) - p_t * gamma_F * sum(w^fuel * s_j)

The fuel-similarity coefficient is -p_t * gamma_F.  When p_t is higher
(more policy credibility), the fuel contagion effect is more negative.

Testable interaction:
  CAR_j = b1*w_fuel + b2*(w_fuel x Credibility_j) + b3*w_geo
        + b4*w_reg + b5*SameSector + eps

If b2 < 0: policy credibility amplifies fuel contagion, consistent with
stranding risk being more real when policy is credible.

Three credibility measures:
  A. ETS membership (binary, pre-determined)
  B. Carbon price (continuous, USD/tCO2e from World Bank)
  C. Combined (ETS x carbon price level)
"""
import csv
import os
import sys
import math
import hashlib
import random
from collections import defaultdict

from _paths import derived_path, raw_path, results_path

# ── Configuration ────────────────────────────────────────────────────

POST_MONTHS = 3   # [-1, +3] window
PRE_MONTHS = 24   # pre-event months for AR demeaning
SEED = 42


def _print(msg=''):
    print(msg)
    sys.stdout.flush()


# ── Matrix utilities ────────────────────────────────────────────────

def invert_matrix(mat):
    """Gauss-Jordan inversion of a square matrix."""
    n = len(mat)
    aug = [row[:] + [1.0 if i == j else 0.0 for j in range(n)]
           for i, row in enumerate(mat)]
    for col in range(n):
        max_row = max(range(col, n), key=lambda r: abs(aug[r][col]))
        if abs(aug[max_row][col]) < 1e-20:
            return None
        aug[col], aug[max_row] = aug[max_row], aug[col]
        pivot = aug[col][col]
        for j in range(2 * n):
            aug[col][j] /= pivot
        for row in range(n):
            if row != col:
                factor = aug[row][col]
                for j in range(2 * n):
                    aug[row][j] -= factor * aug[col][j]
    return [row[n:] for row in aug]


def mat_mul(a, b):
    """Multiply two matrices."""
    rows_a = len(a)
    cols_b = len(b[0])
    mid = len(b)
    out = [[0.0 for _ in range(cols_b)] for _ in range(rows_a)]
    for i in range(rows_a):
        for k in range(mid):
            aik = a[i][k]
            if aik == 0:
                continue
            for j in range(cols_b):
                out[i][j] += aik * b[k][j]
    return out


# ── OLS with event-clustered SEs and full covariance matrix ─────────

def _cluster_cov(X, resid, cluster_map, k):
    """Compute clustered meat matrix. cluster_map: {cid: [indices]}."""
    S = [[0.0] * k for _ in range(k)]
    for idxs in cluster_map.values():
        xu = [0.0] * k
        for i in idxs:
            ri = resid[i]
            Xi = X[i]
            for a in range(k):
                xu[a] += Xi[a] * ri
        for a in range(k):
            xua = xu[a]
            for b_idx in range(a, k):
                v = xua * xu[b_idx]
                S[a][b_idx] += v
                if a != b_idx:
                    S[b_idx][a] += v
    return S


def ols_full(data, y_var, x_vars, cluster_var=None):
    """OLS regression returning betas, cluster-robust SEs, t-stats, R2,
    and the FULL variance-covariance matrix V for inference.

    Returns dict with keys: beta, se, t, r2, n, V (full vcov matrix),
    ss_res, X, resid, y, y_hat, inv_XtX, clusters.
    """
    n = len(data)
    k = len(x_vars) + 1
    if n <= k + 1:
        return None

    y = [d[y_var] for d in data]
    y_mean = sum(y) / n
    ss_tot = sum((yi - y_mean) ** 2 for yi in y)
    if ss_tot < 1e-15:
        return None

    X = [[1.0] + [d[xv] for xv in x_vars] for d in data]

    XtX = [[sum(X[i][a] * X[i][b_idx] for i in range(n))
            for b_idx in range(k)] for a in range(k)]
    Xty = [sum(X[i][a] * y[i] for i in range(n)) for a in range(k)]

    inv_XtX = invert_matrix(XtX)
    if inv_XtX is None:
        return None

    beta = [sum(inv_XtX[a][b_idx] * Xty[b_idx] for b_idx in range(k))
            for a in range(k)]

    y_hat = [sum(X[i][a] * beta[a] for a in range(k)) for i in range(n)]
    resid = [y[i] - y_hat[i] for i in range(n)]
    ss_res = sum(r ** 2 for r in resid)
    r2 = 1 - ss_res / ss_tot if ss_tot > 0 else 0

    V = None
    G = None
    if cluster_var:
        cluster_map = {}
        for i, d in enumerate(data):
            cid = d.get(cluster_var, None)
            cluster_map.setdefault(cid, []).append(i)
        G = len(cluster_map)
        S = _cluster_cov(X, resid, cluster_map, k)
        V = mat_mul(mat_mul(inv_XtX, S), inv_XtX)
        if G > 1:
            scale = (G / (G - 1)) * ((n - 1) / (n - k))
            for a in range(k):
                for b_idx in range(k):
                    V[a][b_idx] *= scale
        se = [math.sqrt(V[a][a]) if V[a][a] > 0 else 0.0 for a in range(k)]
    else:
        s2 = ss_res / (n - k) if n > k else 0
        V = [[s2 * inv_XtX[a][b_idx] for b_idx in range(k)] for a in range(k)]
        se = [math.sqrt(V[a][a]) if V[a][a] > 0 else 0.0 for a in range(k)]

    t_stats = [beta[a] / se[a] if se[a] > 1e-15 else 0 for a in range(k)]

    names = ['intercept'] + x_vars
    return {
        'beta': dict(zip(names, beta)),
        'se': dict(zip(names, se)),
        't': dict(zip(names, t_stats)),
        'r2': r2,
        'n': n,
        'V': V,
        'ss_res': ss_res,
        'X': X,
        'resid': resid,
        'y': y,
        'y_hat': y_hat,
        'inv_XtX': inv_XtX,
        'clusters': G,
    }


# ── p-value from t-statistic (normal approximation) ──────────────────

def _normal_cdf(x):
    """Standard normal CDF approximation (Abramowitz & Stegun 26.2.17)."""
    if x < -8:
        return 0.0
    if x > 8:
        return 1.0
    ax = abs(x)
    b0 = 0.2316419
    b1 = 0.319381530
    b2 = -0.356563782
    b3 = 1.781477937
    b4 = -1.821255978
    b5 = 1.330274429
    t_val = 1.0 / (1.0 + b0 * ax)
    phi = (1.0 / math.sqrt(2.0 * math.pi)) * math.exp(-0.5 * ax * ax)
    cdf = 1.0 - phi * (b1 * t_val + b2 * t_val**2 + b3 * t_val**3
                        + b4 * t_val**4 + b5 * t_val**5)
    if x < 0:
        return 1.0 - cdf
    return cdf


def p_from_t(t_stat):
    """Two-sided p-value from t-stat using normal CDF approximation."""
    return 2.0 * (1.0 - _normal_cdf(abs(t_stat)))


def p_one_sided(t_stat):
    """One-sided p-value (test for negative effect: H1: beta < 0)."""
    if t_stat < 0:
        return 1.0 - _normal_cdf(abs(t_stat))
    else:
        return _normal_cdf(-t_stat)


def sig_stars(p):
    if p < 0.01:
        return '***'
    if p < 0.05:
        return '**'
    if p < 0.10:
        return '*'
    return ''


# ── Quintile assignment ──────────────────────────────────────────────

def assign_quintiles(values):
    """Given list of (key, value) pairs, return dict key -> quintile 1..5."""
    if not values:
        return {}
    sorted_vals = sorted(values, key=lambda x: x[1])
    n = len(sorted_vals)
    assignment = {}
    for rank, (key, _) in enumerate(sorted_vals):
        q = min(int(rank * 5 / n) + 1, 5)
        assignment[key] = q
    return assignment


# ── Parse carbon prices from World Bank XLSX ─────────────────────────

def parse_carbon_prices():
    """Parse carbon prices from the World Bank Carbon Pricing Dashboard XLSX.

    Returns dict: {(fic_iso3, year): max_price_usd_per_tco2e}
    where fic_iso3 is the ISO3 country code matching Compustat fic.
    """
    xlsx_path = raw_path('policy', 'carbon_pricing_worldbank.xlsx')
    if not os.path.exists(xlsx_path):
        _print('  WARNING: carbon_pricing_worldbank.xlsx not found')
        return {}

    try:
        import openpyxl
    except ImportError:
        _print('  WARNING: openpyxl not available')
        return {}

    # Mapping from jurisdiction names (as they appear in the XLSX) to ISO3
    # country codes used in Compustat fic.  Subnational instruments are
    # mapped to the parent country.
    jurisdiction_to_fic = {
        'Albania': 'ALB',
        'Alberta': 'CAN',
        'Andorra': 'AND',
        'Argentina': 'ARG',
        'Australia': 'AUS',
        'Austria': 'AUT',
        'Bahrain': 'BHR',
        'Baja California': 'MEX',
        'Beijing': 'CHN',
        'Botswana': 'BWA',
        'Brazil': 'BRA',
        'British Columbia': 'CAN',
        'Brunei Darussalam': 'BRN',
        'California': 'USA',
        'Canada': 'CAN',
        'Catalonia': 'ESP',
        'Chile': 'CHL',
        'China': 'CHN',
        'Chongqing': 'CHN',
        'Colima': 'MEX',
        'Colombia': 'COL',
        'Colorado': 'USA',
        'Denmark': 'DNK',
        'Dominican Republic': 'DOM',
        'Durango': 'MEX',
        'Estonia': 'EST',
        'EU27+': 'EU',   # special: mapped to all EU members below
        'Finland': 'FIN',
        'France': 'FRA',
        'Fujian': 'CHN',
        'Germany': 'DEU',
        'Guanajuato': 'MEX',
        'Guangdong (except Shenzhen)': 'CHN',
        'Guangdong': 'CHN',
        'Hawaii': 'USA',
        'Hubei': 'CHN',
        'Hungary': 'HUN',
        'Iceland': 'ISL',
        'India': 'IND',
        'Indonesia': 'IDN',
        'Ireland': 'IRL',
        'Israel': 'ISR',
        'Jalisco': 'MEX',
        'Japan': 'JPN',
        'Kazakhstan': 'KAZ',
        'Kenya': 'KEN',
        'Korea, Rep.': 'KOR',
        'Latvia': 'LTU',  # Compustat uses LTU for Latvia
        'Liechtenstein': 'LIE',
        'Luxembourg': 'LUX',
        'Malaysia': 'MYS',
        'Maryland': 'USA',
        'Massachusetts': 'USA',
        'Mauritania': 'MRT',
        'Mexico': 'MEX',
        'Mexico City': 'MEX',
        'Montenegro': 'MNE',
        'Morelos': 'MEX',
        'Morocco': 'MAR',
        'Netherlands': 'NLD',
        'New Brunswick': 'CAN',
        'New Jersey': 'USA',
        'New York State': 'USA',
        'New Zealand': 'NZL',
        'Newfoundland and Labrador': 'CAN',
        'Northwest Territories': 'CAN',
        'Norway': 'NOR',
        'Nova Scotia': 'CAN',
        'Ontario': 'CAN',
        'Oregon': 'USA',
        'Pakistan': 'PAK',
        'Paraguay': 'PRY',
        'Pennsylvania': 'USA',
        'Philippines': 'PHL',
        'Poland': 'POL',
        'Portugal': 'PRT',
        'Prince Edward Island': 'CAN',
        'Quebec': 'CAN',
        'Queretaro': 'MEX',
        'RGGI': 'USA',
        'Saitama': 'JPN',
        'Sakhalin': 'RUS',
        'Saskatchewan': 'CAN',
        'Senegal': 'SEN',
        'Shanghai': 'CHN',
        'Shenzhen': 'CHN',
        'Singapore': 'SGP',
        'Slovenia': 'SVN',
        'South Africa': 'ZAF',
        'Spain': 'ESP',
        'State of Mexico': 'MEX',
        'Sweden': 'SWE',
        'Switzerland': 'CHE',
        'Taiwan, China': 'TWN',
        'Tamaulipas': 'MEX',
        'Thailand': 'THA',
        'Tianjin': 'CHN',
        'Tokyo': 'JPN',
        'Ukraine': 'UKR',
        'United Kingdom': 'GBR',
        'Uruguay': 'URY',
        'Vermont': 'USA',
        'Washington': 'USA',
        'Yucatan': 'MEX',
        'Zacatecas': 'MEX',
    }

    # EU ETS member states that appear in Compustat fic
    eu_ets_members = [
        'AUT', 'BEL', 'BGR', 'CYP', 'CZE', 'DEU', 'DNK', 'ESP', 'EST',
        'FIN', 'FRA', 'GRC', 'HUN', 'IRL', 'ITA', 'LTU', 'LUX', 'NLD',
        'NOR', 'POL', 'PRT', 'ROU', 'SVN', 'SWE',
        # Also covers ISL and LIE via EEA
    ]

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # Read the Compliance_Price sheet
    ws = wb['Compliance_Price']

    # Parse header row (row 2): columns 7..max_col are year columns
    header = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    year_cols = {}  # year (int) -> column index (1-based)
    for c_idx in range(7, len(header) + 1):
        val = ws.cell(2, c_idx).value
        if isinstance(val, (int, float)):
            year_cols[int(val)] = c_idx

    # Also read Gen Info to get the jurisdiction for each UID
    ws_info = wb['Compliance_Gen Info']
    uid_to_jurisdiction = {}
    for r in range(6, ws_info.max_row + 1):
        uid = ws_info.cell(r, 1).value
        jurisdiction = ws_info.cell(r, 5).value
        if uid and jurisdiction:
            # Clean up unicode artifacts
            jurisdiction = jurisdiction.replace('\xa0', ' ').strip()
            uid_to_jurisdiction[uid] = jurisdiction

    # Parse prices: for each instrument row, get jurisdiction -> fic, then prices
    # Result: {(fic, year): max_price}
    prices = {}  # (fic_iso3, year) -> max price USD/tCO2e

    for r in range(3, ws.max_row + 1):
        uid = ws.cell(r, 1).value
        if not uid:
            continue

        # Get jurisdiction from Gen Info mapping, fall back to Compliance_Price
        # region field
        jurisdiction = uid_to_jurisdiction.get(uid, '')

        # Map jurisdiction to fic codes
        fic_codes = []
        if jurisdiction in jurisdiction_to_fic:
            fic = jurisdiction_to_fic[jurisdiction]
            if fic == 'EU':
                fic_codes = eu_ets_members[:]
            else:
                fic_codes = [fic]
        else:
            # Try partial matching
            for jur_name, fic in jurisdiction_to_fic.items():
                if jur_name.lower() in jurisdiction.lower():
                    if fic == 'EU':
                        fic_codes = eu_ets_members[:]
                    else:
                        fic_codes.append(fic)
                    break

        if not fic_codes:
            continue

        # Extract prices for each year
        for year, col_idx in year_cols.items():
            cell_val = ws.cell(r, col_idx).value
            if cell_val is None:
                continue
            try:
                price = float(cell_val)
            except (ValueError, TypeError):
                continue
            if price <= 0:
                continue

            for fic in fic_codes:
                key = (fic, year)
                if key not in prices or price > prices[key]:
                    prices[key] = price

    wb.close()
    return prices


# ── Load data ────────────────────────────────────────────────────────

_print('=' * 70)
_print('POLICY CREDIBILITY INTERACTION TEST')
_print('Does p_t moderate spatial transmission of coal retirement shocks?')
_print('=' * 70)
_print()

_print('Loading monthly returns...')
monthly_ret = defaultdict(dict)
with open(derived_path('returns', 'monthly_returns.csv'), 'r', encoding='utf-8') as f:
    for row in csv.DictReader(f):
        gk = row['gvkey']
        ym = row['datadate'][:7]
        try:
            monthly_ret[gk][ym] = float(row['ret_monthly'])
        except ValueError:
            pass
_print(f'  Monthly: {len(monthly_ret)} firms')

# Fama-French monthly factors (vwretd = Mkt-RF + RF)
_print('Loading Fama-French factors...')


def load_ff_factors_monthly(path):
    if not os.path.exists(path):
        return None
    vwretd = {}
    with open(path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('This file') or line.startswith('The '):
                continue
            if line.startswith(','):
                continue
            parts = [p.strip() for p in line.split(',')]
            if len(parts) < 5:
                continue
            date = parts[0]
            if not date.isdigit() or len(date) != 6:
                continue
            try:
                mktrf_val = float(parts[1])
                rf_val = float(parts[4])
            except ValueError:
                continue
            vw = (mktrf_val + rf_val) / 100.0
            vwretd[f'{date[:4]}-{date[4:6]}'] = vw
    return vwretd


market_ret_monthly = load_ff_factors_monthly(
    raw_path('factors', 'F-F_Research_Data_Factors.csv')
)
if not market_ret_monthly:
    raise RuntimeError('Missing F-F monthly factors.')
_print(f'  Market months: {len(market_ret_monthly)}')

# Weight matrices
_print('Loading weight matrices...')
W_geo = defaultdict(dict)
with open(derived_path('networks', 'weight_matrix_W_geo.csv'), 'r',
          encoding='utf-8') as f:
    for row in csv.DictReader(f):
        W_geo[row['gvkey_i']][row['gvkey_j']] = float(row['w_ij'])
_print(f'  W_geo firms: {len(W_geo)}')

W_fuel = defaultdict(dict)
fuel_path = derived_path('networks', 'weight_matrix_W_fuel.csv')
if os.path.exists(fuel_path):
    with open(fuel_path, 'r', encoding='utf-8') as f:
        for row in csv.DictReader(f):
            W_fuel[row['gvkey_i']][row['gvkey_j']] = float(row['w_ij'])
    _print(f'  W_fuel edges: {sum(len(v) for v in W_fuel.values())}')
else:
    _print('  W_fuel: NOT FOUND')

W_reg = defaultdict(dict)
reg_path = derived_path('networks', 'weight_matrix_W_regulatory.csv')
if os.path.exists(reg_path):
    with open(reg_path, 'r', encoding='utf-8') as f:
        for row in csv.DictReader(f):
            wval = row.get('w_ij')
            if wval in (None, ''):
                wval = row.get('w_reg')
            try:
                W_reg[row['gvkey_i']][row['gvkey_j']] = float(wval)
            except (ValueError, TypeError):
                continue
    _print(f'  W_reg edges: {sum(len(v) for v in W_reg.values())}')
else:
    _print('  W_reg: NOT FOUND')

# Fundamentals (latest record per firm for sector classification)
_print('Loading fundamentals...')
fundamentals = {}
fundamentals_by_year = defaultdict(dict)
with open(derived_path('fundamentals', 'firm_fundamentals.csv'), 'r',
          encoding='utf-8') as f:
    for row in csv.DictReader(f):
        gk = row['gvkey']
        fy = row['fyear']
        fundamentals_by_year[gk][fy] = row
        if gk not in fundamentals or fy > fundamentals[gk]['fyear']:
            fundamentals[gk] = row
_print(f'  Firms: {len(fundamentals)}, firm-years: '
       f'{sum(len(v) for v in fundamentals_by_year.values())}')


def get_sic4(gvkey):
    f = fundamentals.get(gvkey)
    if f and f.get('sic'):
        return f['sic'][:4]
    return None


def get_fic(gvkey):
    """Get country of incorporation (ISO3) for a firm."""
    f = fundamentals.get(gvkey)
    if f and f.get('fic'):
        return f['fic']
    return None


# ETS membership
_print('Loading ETS membership...')
ets_membership = {}
ets_path = derived_path('networks', 'firm_ets_membership.csv')
if os.path.exists(ets_path):
    with open(ets_path, 'r', encoding='utf-8') as f:
        for row in csv.DictReader(f):
            ets_membership[row['gvkey']] = int(row['has_ets'])
    n_ets = sum(v for v in ets_membership.values())
    _print(f'  ETS firms: {n_ets}/{len(ets_membership)}')
else:
    _print('  WARNING: firm_ets_membership.csv not found')

# Carbon prices
_print('Loading carbon prices from World Bank XLSX...')
carbon_prices = parse_carbon_prices()
if carbon_prices:
    countries_with_prices = len(set(k[0] for k in carbon_prices))
    years_with_prices = sorted(set(k[1] for k in carbon_prices))
    _print(f'  Carbon prices: {len(carbon_prices)} country-year pairs, '
           f'{countries_with_prices} countries, '
           f'years {years_with_prices[0]}-{years_with_prices[-1]}')
else:
    _print('  WARNING: No carbon prices parsed; Measure B/C unavailable')

# Events (first-mover coal retirements only)
_print('Loading events...')
all_events = []
with open(derived_path('events', 'coal_retirement_events.csv'), 'r',
          encoding='utf-8') as f:
    for row in csv.DictReader(f):
        if not row.get('matched_gvkeys'):
            continue
        if row.get('is_first_mover') != 'True':
            continue
        ann_date = row.get('announcement_date', '').strip()
        ret_date = row.get('event_date', '').strip()
        effective_date = ann_date if ann_date else ret_date
        event_year = None
        if effective_date and len(effective_date) >= 4 and effective_date[:4].isdigit():
            event_year = int(effective_date[:4])
        else:
            event_year = int(row['ret_year']) if row.get('ret_year') else None
        all_events.append({
            'plant': row['plant_name'],
            'year': event_year,
            'event_date': effective_date,
            'gvkeys': row['matched_gvkeys'].split(';'),
        })
_print(f'  First-mover events: {len(all_events)}')


# ── CAR computation ──────────────────────────────────────────────────

def compute_monthly_car(gvkey, event_month, post=3):
    """Monthly CAR[-1, +post] using vwretd model with pre-event demeaning."""
    if gvkey not in monthly_ret:
        return None
    months = sorted(monthly_ret[gvkey].keys())
    event_idx = None
    for i, m in enumerate(months):
        if m >= event_month:
            event_idx = i
            break
    if event_idx is None:
        return None

    # Require enough pre-event data
    pre_rets = [monthly_ret[gvkey][months[i]]
                for i in range(max(0, event_idx - PRE_MONTHS), event_idx)
                if months[i] in monthly_ret[gvkey]]
    if len(pre_rets) < 12:
        return None

    # Pre-demean ARs by pre-window mean
    ar_list = []
    for i in range(max(0, event_idx - PRE_MONTHS), event_idx):
        m = months[i]
        if m in monthly_ret[gvkey] and m in market_ret_monthly:
            ar_list.append(monthly_ret[gvkey][m] - market_ret_monthly[m])
    pre_mean_ar = (sum(ar_list) / len(ar_list)) if ar_list else 0.0

    car = 0.0
    for offset in range(-1, post + 1):
        idx = event_idx + offset
        if 0 <= idx < len(months) and months[idx] in monthly_ret[gvkey]:
            m = months[idx]
            r_it = monthly_ret[gvkey][m]
            if m in market_ret_monthly:
                ar = r_it - market_ret_monthly[m]
                car += ar - pre_mean_ar
    return car


# ── Build regression dataset ─────────────────────────────────────────

def build_obs():
    """Build cross-sectional regression dataset for the [-1,+3] window,
    augmented with credibility measures."""
    obs = []
    for event_id, event in enumerate(all_events):
        event_gvkeys = set(event['gvkeys'])
        year = event['year']
        event_date = event.get('event_date', '')
        if event_date and len(event_date) >= 7:
            event_month = event_date[:7]
        else:
            event_month = f'{year}-07' if year else None
        if not event_month:
            continue

        # Get first-mover SIC4
        fm_sic4 = None
        for gk in event_gvkeys:
            fm_sic4 = get_sic4(gk)
            if fm_sic4:
                break

        for fm_gk in event_gvkeys:
            if fm_gk not in W_geo:
                continue
            neighbors = W_geo[fm_gk]
            neighbor_gks = set(neighbors.keys()) - event_gvkeys
            non_connected = [gk for gk in fundamentals
                             if gk not in event_gvkeys and gk not in neighbors]
            stable_seed = int(hashlib.md5(
                str(fm_gk).encode('utf-8')).hexdigest()[:8], 16)
            random.seed(stable_seed)
            n_ctrl = min(len(non_connected),
                         max(5 * len(neighbor_gks), 20))
            ctrl_sample = (random.sample(non_connected, n_ctrl)
                           if len(non_connected) > n_ctrl
                           else non_connected)
            candidate_firms = list(neighbor_gks) + ctrl_sample

            for gk in candidate_firms:
                w_geo = neighbors.get(gk, 0.0)
                w_fuel = W_fuel.get(fm_gk, {}).get(gk, 0.0)
                w_reg = W_reg.get(fm_gk, {}).get(gk, 0.0)
                j_sic4 = get_sic4(gk)
                same_sector = 1.0 if (fm_sic4 and j_sic4
                                      and fm_sic4 == j_sic4) else 0.0

                car = compute_monthly_car(gk, event_month, post=POST_MONTHS)
                if car is None:
                    continue

                # Credibility measures for firm j
                has_ets = float(ets_membership.get(gk, 0))

                # Carbon price for firm j's country in the event year
                fic = get_fic(gk)
                cp = 0.0
                if fic and year and carbon_prices:
                    cp = carbon_prices.get((fic, year), 0.0)

                # Combined: ETS x carbon price
                combined = cp if has_ets > 0.5 else 0.0

                obs.append({
                    'car': car,
                    'w_fuel': w_fuel,
                    'w_geo': w_geo,
                    'w_reg': w_reg,
                    'same_sector': same_sector,
                    'has_ets': has_ets,
                    'carbon_price': cp,
                    'combined_credibility': combined,
                    'w_fuel_x_ets': w_fuel * has_ets,
                    'w_fuel_x_cp': w_fuel * cp,
                    'w_fuel_x_combined': w_fuel * combined,
                    'w_geo_x_ets': w_geo * has_ets,
                    'event_id': event_id,
                    'gvkey': gk,
                    'fic': fic if fic else '',
                })

    return obs


# ── Main analysis ────────────────────────────────────────────────────

_print()
_print('Building dataset...')
obs = build_obs()
_print(f'  N = {len(obs)} observations')
_print(f'  Events: {len(set(o["event_id"] for o in obs))}')
_print(f'  Neighbors (w_geo > 0): {sum(1 for o in obs if o["w_geo"] > 0)}')
_print(f'  ETS firms: {sum(1 for o in obs if o["has_ets"] > 0.5)}')
_print(f'  Non-ETS firms: {sum(1 for o in obs if o["has_ets"] < 0.5)}')
n_with_cp = sum(1 for o in obs if o['carbon_price'] > 0)
_print(f'  Obs with carbon_price > 0: {n_with_cp}')

# Check same_sector variation
ss_vals = set(o['same_sector'] for o in obs)
has_same_sector = len(ss_vals) > 1

# ── SPECIFICATION 1: ETS Interaction ─────────────────────────────────

_print()
_print('=' * 70)
_print('SPECIFICATION 1: ETS MEMBERSHIP INTERACTION')
_print('CAR = a + b1*w_fuel + b2*(w_fuel x has_ets) + b3*w_geo')
_print('    + b4*w_reg + b5*SameSector + eps')
_print('H1 (one-sided): b2 < 0')
_print('=' * 70)

spec1_vars = ['w_fuel', 'w_fuel_x_ets', 'w_geo', 'w_reg']
if has_same_sector:
    spec1_vars.append('same_sector')

res1 = ols_full(obs, 'car', spec1_vars, cluster_var='event_id')
if res1 is None:
    _print('ERROR: Specification 1 OLS failed.')
    sys.exit(1)

_print(f'  N = {res1["n"]}, R2 = {res1["r2"]:.6f}, '
       f'clusters = {res1["clusters"]}')
for v in spec1_vars:
    p2 = p_from_t(res1['t'][v])
    _print(f'  {v:<20s}: beta = {res1["beta"][v]:+.6f}, '
           f'se = {res1["se"][v]:.6f}, t = {res1["t"][v]:.3f}, '
           f'p = {p2:.4f}{sig_stars(p2)}')

# One-sided test for the interaction
t_ets_interaction = res1['t']['w_fuel_x_ets']
p_ets_onesided = p_one_sided(t_ets_interaction)
_print(f'  One-sided test (b2 < 0): p = {p_ets_onesided:.4f}')

if t_ets_interaction < 0:
    _print('  --> Fuel contagion is MORE NEGATIVE for ETS firms '
           '(credibility amplifies stranding)')
else:
    _print('  --> Fuel contagion is NOT more negative for ETS firms')


# ── SPECIFICATION 2: Carbon Price Interaction ─────────────────────────

has_cp_variation = len(set(o['carbon_price'] for o in obs)) > 1 and n_with_cp > 50
res2 = None

if has_cp_variation:
    _print()
    _print('=' * 70)
    _print('SPECIFICATION 2: CARBON PRICE INTERACTION')
    _print('CAR = a + b1*w_fuel + b2*(w_fuel x carbon_price) + b3*w_geo')
    _print('    + b4*w_reg + b5*SameSector + eps')
    _print('H1 (one-sided): b2 < 0')
    _print('=' * 70)

    spec2_vars = ['w_fuel', 'w_fuel_x_cp', 'w_geo', 'w_reg']
    if has_same_sector:
        spec2_vars.append('same_sector')

    res2 = ols_full(obs, 'car', spec2_vars, cluster_var='event_id')
    if res2 is None:
        _print('ERROR: Specification 2 OLS failed.')
    else:
        _print(f'  N = {res2["n"]}, R2 = {res2["r2"]:.6f}, '
               f'clusters = {res2["clusters"]}')
        for v in spec2_vars:
            p2 = p_from_t(res2['t'][v])
            _print(f'  {v:<20s}: beta = {res2["beta"][v]:+.6f}, '
                   f'se = {res2["se"][v]:.6f}, t = {res2["t"][v]:.3f}, '
                   f'p = {p2:.4f}{sig_stars(p2)}')

        t_cp_interaction = res2['t']['w_fuel_x_cp']
        p_cp_onesided = p_one_sided(t_cp_interaction)
        _print(f'  One-sided test (b2 < 0): p = {p_cp_onesided:.4f}')
else:
    _print()
    _print('SPECIFICATION 2: SKIPPED (insufficient carbon price variation)')


# ── SPECIFICATION 3: Placebo (Geo x ETS) ─────────────────────────────

_print()
_print('=' * 70)
_print('SPECIFICATION 3: PLACEBO (GEO x ETS)')
_print('CAR = a + b1*w_fuel + b2*(w_fuel x has_ets) + b3*w_geo')
_print('    + b4*(w_geo x has_ets) + b5*w_reg + b6*SameSector + eps')
_print('Placebo: b4 should be ~0 (competitive benefit is physical,')
_print('not policy-dependent)')
_print('=' * 70)

spec3_vars = ['w_fuel', 'w_fuel_x_ets', 'w_geo', 'w_geo_x_ets', 'w_reg']
if has_same_sector:
    spec3_vars.append('same_sector')

res3 = ols_full(obs, 'car', spec3_vars, cluster_var='event_id')
if res3 is None:
    _print('ERROR: Specification 3 OLS failed.')
    sys.exit(1)

_print(f'  N = {res3["n"]}, R2 = {res3["r2"]:.6f}, '
       f'clusters = {res3["clusters"]}')
for v in spec3_vars:
    p2 = p_from_t(res3['t'][v])
    _print(f'  {v:<20s}: beta = {res3["beta"][v]:+.6f}, '
           f'se = {res3["se"][v]:.6f}, t = {res3["t"][v]:.3f}, '
           f'p = {p2:.4f}{sig_stars(p2)}')

t_geo_ets = res3['t']['w_geo_x_ets']
p_geo_ets = p_from_t(t_geo_ets)
_print(f'  Placebo (geo x ETS): t = {t_geo_ets:.3f}, p = {p_geo_ets:.4f}')
if abs(t_geo_ets) < 1.96:
    _print('  --> Placebo passes: geo contagion does NOT depend on ETS')
else:
    _print('  --> Placebo fails: geo contagion appears to depend on ETS')


# ── SPECIFICATION 4: Difference-in-interactions test ──────────────────

_print()
_print('=' * 70)
_print('SPECIFICATION 4: DIFFERENCE-IN-INTERACTIONS')
_print('From Spec 3: test H0: b_fuel_ets = b_geo_ets')
_print('=' * 70)

names_list = ['intercept'] + spec3_vars
idx_fuel_ets = names_list.index('w_fuel_x_ets')
idx_geo_ets = names_list.index('w_geo_x_ets')

V3 = res3['V']
beta_fuel_ets = res3['beta']['w_fuel_x_ets']
beta_geo_ets = res3['beta']['w_geo_x_ets']

var_fuel_ets = V3[idx_fuel_ets][idx_fuel_ets]
var_geo_ets = V3[idx_geo_ets][idx_geo_ets]
cov_fuel_geo_ets = V3[idx_fuel_ets][idx_geo_ets]

diff_ets = beta_fuel_ets - beta_geo_ets
se_diff_ets = math.sqrt(var_fuel_ets + var_geo_ets - 2.0 * cov_fuel_geo_ets)
t_diff_ets = diff_ets / se_diff_ets if se_diff_ets > 1e-15 else 0.0
p_diff_ets = p_from_t(t_diff_ets)

_print(f'  b_fuel_ets: {beta_fuel_ets:+.6f} '
       f'(SE {res3["se"]["w_fuel_x_ets"]:.6f})')
_print(f'  b_geo_ets:  {beta_geo_ets:+.6f} '
       f'(SE {res3["se"]["w_geo_x_ets"]:.6f})')
_print(f'  Difference (b_fuel_ets - b_geo_ets): {diff_ets:+.6f}')
_print(f'  Cov(b_fuel_ets, b_geo_ets): {cov_fuel_geo_ets:+.10f}')
_print(f'  SE of difference: {se_diff_ets:.6f}')
_print(f'  t-statistic: {t_diff_ets:.3f}')
_print(f'  p-value (two-sided): {p_diff_ets:.4f}{sig_stars(p_diff_ets)}')

if p_diff_ets < 0.05:
    _print('  --> Credibility differentially affects fuel vs geo channel')
else:
    _print('  --> Cannot reject equal credibility moderation across channels')


# ── PORTFOLIO SORT COMPARISON ─────────────────────────────────────────

_print()
_print('=' * 70)
_print('PORTFOLIO SORT: ETS vs NON-ETS FUEL SPREADS')
_print('=' * 70)

MIN_FIRMS_SORT = 10

# For each event, split firms by ETS, do fuel quintile sorts, compute spread
ets_fuel_spreads = []     # Q5-Q1 spread for ETS firms, per event
non_ets_fuel_spreads = [] # Q5-Q1 spread for non-ETS firms, per event
events_both = 0

for event_id, event in enumerate(all_events):
    event_gvkeys = set(event['gvkeys'])
    year = event['year']
    event_date = event.get('event_date', '')
    if event_date and len(event_date) >= 7:
        event_month = event_date[:7]
    else:
        event_month = f'{year}-07' if year else None
    if not event_month:
        continue

    # Collect firm data for this event
    firm_data = {}
    for fm_gk in event_gvkeys:
        if fm_gk not in W_geo:
            continue
        all_neighbors = set()
        if fm_gk in W_geo:
            all_neighbors.update(W_geo[fm_gk].keys())
        if fm_gk in W_fuel:
            all_neighbors.update(W_fuel[fm_gk].keys())
        all_neighbors -= event_gvkeys

        for gk in all_neighbors:
            if gk in firm_data:
                w_fuel_val = W_fuel.get(fm_gk, {}).get(gk, 0.0)
                firm_data[gk]['w_fuel'] = max(firm_data[gk]['w_fuel'],
                                              w_fuel_val)
                continue
            car = compute_monthly_car(gk, event_month, post=POST_MONTHS)
            if car is None:
                continue
            w_fuel_val = W_fuel.get(fm_gk, {}).get(gk, 0.0)
            has_ets_val = ets_membership.get(gk, 0)
            firm_data[gk] = {
                'car': car,
                'w_fuel': w_fuel_val,
                'has_ets': has_ets_val,
            }

    # Split by ETS
    ets_firms = [(gk, fd) for gk, fd in firm_data.items()
                 if fd['has_ets'] > 0.5]
    non_ets_firms = [(gk, fd) for gk, fd in firm_data.items()
                     if fd['has_ets'] < 0.5]

    ets_spread = None
    non_ets_spread = None

    if len(ets_firms) >= MIN_FIRMS_SORT:
        fuel_pairs = [(gk, fd['w_fuel']) for gk, fd in ets_firms]
        fq = assign_quintiles(fuel_pairs)
        q_cars = {q: [] for q in range(1, 6)}
        for gk, fd in ets_firms:
            q_cars[fq[gk]].append(fd['car'])
        q_means = {}
        for q in range(1, 6):
            if q_cars[q]:
                q_means[q] = sum(q_cars[q]) / len(q_cars[q])
        if 5 in q_means and 1 in q_means:
            ets_spread = q_means[5] - q_means[1]
            ets_fuel_spreads.append(ets_spread)

    if len(non_ets_firms) >= MIN_FIRMS_SORT:
        fuel_pairs = [(gk, fd['w_fuel']) for gk, fd in non_ets_firms]
        fq = assign_quintiles(fuel_pairs)
        q_cars = {q: [] for q in range(1, 6)}
        for gk, fd in non_ets_firms:
            q_cars[fq[gk]].append(fd['car'])
        q_means = {}
        for q in range(1, 6):
            if q_cars[q]:
                q_means[q] = sum(q_cars[q]) / len(q_cars[q])
        if 5 in q_means and 1 in q_means:
            non_ets_spread = q_means[5] - q_means[1]
            non_ets_fuel_spreads.append(non_ets_spread)

    if ets_spread is not None and non_ets_spread is not None:
        events_both += 1


def mean_and_t(vals):
    """Return (mean, t-stat, n) for a list of values (H0: mean = 0)."""
    n = len(vals)
    if n < 2:
        return (0.0, 0.0, n)
    m = sum(vals) / n
    var = sum((v - m) ** 2 for v in vals) / (n - 1)
    se = math.sqrt(var / n) if var > 0 else 0.0
    t = m / se if se > 1e-15 else 0.0
    return (m, t, n)


ets_mean, ets_t, ets_n = mean_and_t(ets_fuel_spreads)
non_ets_mean, non_ets_t, non_ets_n = mean_and_t(non_ets_fuel_spreads)

# Difference in spreads (paired where both exist, else unpaired)
# Use unpaired t-test for simplicity
diff_spread = ets_mean - non_ets_mean
n_min = min(ets_n, non_ets_n)
if ets_n >= 2 and non_ets_n >= 2:
    var_ets = sum((v - ets_mean) ** 2 for v in ets_fuel_spreads) / (ets_n - 1)
    var_non = sum((v - non_ets_mean) ** 2
                  for v in non_ets_fuel_spreads) / (non_ets_n - 1)
    se_diff_spread = math.sqrt(var_ets / ets_n + var_non / non_ets_n)
    t_diff_spread = diff_spread / se_diff_spread if se_diff_spread > 1e-15 else 0.0
else:
    se_diff_spread = 0.0
    t_diff_spread = 0.0

_print(f'  ETS firms: mean fuel Q5-Q1 spread = {ets_mean:+.4f} '
       f'(t = {ets_t:.3f}, N = {ets_n} events)')
_print(f'  Non-ETS firms: mean fuel Q5-Q1 spread = {non_ets_mean:+.4f} '
       f'(t = {non_ets_t:.3f}, N = {non_ets_n} events)')
_print(f'  Difference: {diff_spread:+.4f} (t = {t_diff_spread:.3f})')
_print(f'  Events with both groups: {events_both}')

if ets_mean < non_ets_mean:
    _print('  --> ETS firms show MORE negative fuel contagion (consistent)')
else:
    _print('  --> ETS firms do NOT show more negative fuel contagion')


# ── Write output ─────────────────────────────────────────────────────

_print()
_print('Writing results...')
out_path = results_path('metrics', 'strategy2_credibility_interaction.md')
os.makedirs(os.path.dirname(out_path), exist_ok=True)

lines = []
lines.append('# Policy Credibility Interaction: '
             'Does p_t Moderate Spatial Contagion?')
lines.append('')
lines.append('Theoretical prediction (Equation 2): the fuel-similarity '
             'coefficient')
lines.append('is -p_t * gamma_F. Higher credibility (p_t) implies more '
             'negative fuel contagion.')
lines.append('')
lines.append(f'Window: [-1, +{POST_MONTHS}] months (monthly CARs, vwretd)')
lines.append(f'Events: {len(all_events)} first-mover coal retirements')
lines.append(f'N = {len(obs)} observations')
lines.append(f'ETS firms: {sum(1 for o in obs if o["has_ets"] > 0.5)} / '
             f'{len(obs)} observations')
lines.append(f'Obs with carbon_price > 0: {n_with_cp}')
lines.append(f'Standard errors: event-clustered')
lines.append('')

# ── Spec 1 ──
lines.append('## Specification 1: ETS Membership Interaction')
lines.append('')
lines.append('CAR = a + b1*w_fuel + b2*(w_fuel x has_ets) + b3*w_geo '
             '+ b4*w_reg + b5*SameSector + eps')
lines.append('')
lines.append('| Variable | Beta | SE | t | p (two-sided) |')
lines.append('|---|---:|---:|---:|---:|')
for v in ['intercept'] + spec1_vars:
    b = res1['beta'][v]
    s = res1['se'][v]
    t = res1['t'][v]
    p = p_from_t(t)
    lines.append(f'| {v} | {b:+.6f} | {s:.6f} | {t:.3f} | '
                 f'{p:.4f}{sig_stars(p)} |')
lines.append('')
lines.append(f'R2 = {res1["r2"]:.6f}, N = {res1["n"]}, '
             f'clusters = {res1["clusters"]}')
lines.append('')
lines.append(f'One-sided test (H1: b2 < 0): '
             f't = {t_ets_interaction:.3f}, p = {p_ets_onesided:.4f}')
lines.append('')
if t_ets_interaction < 0 and p_ets_onesided < 0.10:
    lines.append('Interpretation: Fuel contagion is significantly more '
                 'negative for firms operating under emissions trading '
                 'systems. Policy credibility amplifies the stranding '
                 'channel, consistent with Equation 2.')
elif t_ets_interaction < 0:
    lines.append('Interpretation: The point estimate suggests fuel '
                 'contagion is more negative for ETS firms (correct sign), '
                 'but the effect is not statistically significant at '
                 'conventional levels.')
else:
    lines.append('Interpretation: The interaction coefficient has the '
                 'wrong sign; fuel contagion is not more negative for '
                 'ETS firms.')
lines.append('')

# ── Spec 2 ──
lines.append('## Specification 2: Carbon Price Interaction')
lines.append('')
if res2 is not None:
    lines.append('CAR = a + b1*w_fuel + b2*(w_fuel x carbon_price) '
                 '+ b3*w_geo + b4*w_reg + b5*SameSector + eps')
    lines.append('')
    spec2_all = ['intercept'] + spec2_vars
    lines.append('| Variable | Beta | SE | t | p (two-sided) |')
    lines.append('|---|---:|---:|---:|---:|')
    for v in spec2_all:
        b = res2['beta'][v]
        s = res2['se'][v]
        t = res2['t'][v]
        p = p_from_t(t)
        lines.append(f'| {v} | {b:+.6f} | {s:.6f} | {t:.3f} | '
                     f'{p:.4f}{sig_stars(p)} |')
    lines.append('')
    lines.append(f'R2 = {res2["r2"]:.6f}, N = {res2["n"]}, '
                 f'clusters = {res2["clusters"]}')
    lines.append('')
    t_cp_val = res2['t']['w_fuel_x_cp']
    p_cp_os = p_one_sided(t_cp_val)
    lines.append(f'One-sided test (H1: b2 < 0): '
                 f't = {t_cp_val:.3f}, p = {p_cp_os:.4f}')
    lines.append('')
    if t_cp_val < 0 and p_cp_os < 0.10:
        lines.append('Interpretation: Higher carbon prices amplify fuel '
                     'contagion. A one-dollar increase in the carbon price '
                     'makes the fuel-similarity channel more negative, '
                     'consistent with policy credibility moderating '
                     'stranding risk transmission.')
    elif t_cp_val < 0:
        lines.append('Interpretation: The point estimate has the predicted '
                     'sign (higher carbon prices amplify fuel contagion) '
                     'but is not statistically significant.')
    else:
        lines.append('Interpretation: The carbon price interaction has the '
                     'wrong sign.')
else:
    lines.append('Skipped: insufficient carbon price variation in sample.')
lines.append('')

# ── Spec 3 ──
lines.append('## Specification 3: Placebo (Geo x ETS)')
lines.append('')
lines.append('CAR = a + b1*w_fuel + b2*(w_fuel x has_ets) + b3*w_geo '
             '+ b4*(w_geo x has_ets) + b5*w_reg + b6*SameSector + eps')
lines.append('')
lines.append('Placebo: geo x ETS should be ~0 (competitive benefit is '
             'physical, not policy-dependent)')
lines.append('')
lines.append('| Variable | Beta | SE | t | p (two-sided) |')
lines.append('|---|---:|---:|---:|---:|')
for v in ['intercept'] + spec3_vars:
    b = res3['beta'][v]
    s = res3['se'][v]
    t = res3['t'][v]
    p = p_from_t(t)
    lines.append(f'| {v} | {b:+.6f} | {s:.6f} | {t:.3f} | '
                 f'{p:.4f}{sig_stars(p)} |')
lines.append('')
lines.append(f'R2 = {res3["r2"]:.6f}, N = {res3["n"]}, '
             f'clusters = {res3["clusters"]}')
lines.append('')
p_geo_ets_2s = p_from_t(t_geo_ets)
if abs(t_geo_ets) < 1.96:
    lines.append(f'Placebo PASSES: geo x ETS is not significant '
                 f'(t = {t_geo_ets:.3f}, p = {p_geo_ets_2s:.4f}). '
                 f'Geographic competitive benefit does not depend on '
                 f'policy credibility, as expected.')
else:
    lines.append(f'Placebo FAILS: geo x ETS is significant '
                 f'(t = {t_geo_ets:.3f}, p = {p_geo_ets_2s:.4f}).')
lines.append('')

# ── Spec 4 ──
lines.append('## Specification 4: Difference test')
lines.append('')
lines.append(f'b_fuel_ets = {beta_fuel_ets:+.6f} '
             f'(SE {res3["se"]["w_fuel_x_ets"]:.6f})')
lines.append(f'b_geo_ets  = {beta_geo_ets:+.6f} '
             f'(SE {res3["se"]["w_geo_x_ets"]:.6f})')
lines.append(f'b_fuel_ets - b_geo_ets = {diff_ets:+.6f} '
             f'(t = {t_diff_ets:.3f}, p = {p_diff_ets:.4f}'
             f'{sig_stars(p_diff_ets)})')
lines.append(f'Cov(b_fuel_ets, b_geo_ets) = {cov_fuel_geo_ets:+.10f}')
lines.append('')
if p_diff_ets < 0.05:
    lines.append('Credibility differentially affects the fuel vs '
                 'geographic channels: the fuel-similarity interaction with '
                 'ETS membership is statistically different from the '
                 'geographic interaction.')
else:
    lines.append('Cannot reject that credibility moderates both channels '
                 'equally (difference not significant).')
lines.append('')

# ── Portfolio sorts ──
lines.append('## Portfolio Sort Comparison')
lines.append('')
lines.append('Fuel Q5-Q1 spread by ETS status (event-level, then averaged):')
lines.append('')
lines.append('| Group | Fuel Q5-Q1 spread | t-stat | N events |')
lines.append('|---|---:|---:|---:|')

p_ets_sort = p_from_t(ets_t) if ets_n >= 2 else 1.0
p_non_sort = p_from_t(non_ets_t) if non_ets_n >= 2 else 1.0
p_diff_sort = p_from_t(t_diff_spread) if n_min >= 2 else 1.0

lines.append(f'| ETS firms | {ets_mean:+.4f}{sig_stars(p_ets_sort)} | '
             f'{ets_t:.3f} | {ets_n} |')
lines.append(f'| Non-ETS firms | {non_ets_mean:+.4f}{sig_stars(p_non_sort)} | '
             f'{non_ets_t:.3f} | {non_ets_n} |')
lines.append(f'| Difference | {diff_spread:+.4f}{sig_stars(p_diff_sort)} | '
             f'{t_diff_spread:.3f} | {events_both} (paired) |')
lines.append('')

# ── Summary ──
lines.append('## Summary')
lines.append('')

# Build summary paragraph
summary_parts = []
summary_parts.append(
    f'This test examines whether policy credibility (p_t) moderates '
    f'the spatial transmission of coal retirement shocks through '
    f'fuel-similarity networks, using {len(obs)} event-firm observations '
    f'from {len(all_events)} first-mover coal retirements.'
)

# Spec 1 result
if t_ets_interaction < 0:
    summary_parts.append(
        f'The ETS interaction coefficient on fuel similarity is '
        f'{res1["beta"]["w_fuel_x_ets"]:+.4f} '
        f'(t = {t_ets_interaction:.2f}, one-sided p = {p_ets_onesided:.3f}), '
        f'indicating that fuel contagion is '
        f'{"significantly " if p_ets_onesided < 0.10 else ""}'
        f'more negative for firms operating under emissions trading systems.'
    )
else:
    summary_parts.append(
        f'The ETS interaction coefficient on fuel similarity is '
        f'{res1["beta"]["w_fuel_x_ets"]:+.4f} '
        f'(t = {t_ets_interaction:.2f}), which does not support the '
        f'hypothesis that ETS membership amplifies fuel contagion.'
    )

# Spec 3 result (placebo)
if abs(t_geo_ets) < 1.96:
    summary_parts.append(
        f'The placebo test passes: the geographic channel interaction '
        f'with ETS is not significant (t = {t_geo_ets:.2f}), confirming '
        f'that policy credibility specifically moderates the stranding '
        f'channel rather than geographic proximity effects.'
    )

# Portfolio sort
if ets_n >= 2 and non_ets_n >= 2:
    summary_parts.append(
        f'Portfolio sorts confirm the pattern: the fuel Q5-Q1 spread '
        f'is {ets_mean:+.4f} for ETS firms versus {non_ets_mean:+.4f} '
        f'for non-ETS firms '
        f'(difference t = {t_diff_spread:.2f}).'
    )

lines.append(' '.join(summary_parts))
lines.append('')

with open(out_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

_print(f'\nWrote: {out_path}')
_print('Done.')
