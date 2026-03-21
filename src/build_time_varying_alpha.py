"""Build time-varying alpha (fossil intensity) for each firm-year.

For each matched Compustat firm, reconstructs the operating fleet in each year
using GEM plant start/retirement years. Computes:
  alpha_it = fossil_MW_operating_in_year_t / total_MW_operating_in_year_t

Output: firm_alpha_panel.csv with columns:
  gvkey, year, coal_mw, gas_mw, fossil_mw, clean_mw, total_mw,
  alpha (fossil share), coal_share, n_plants
"""
import openpyxl, csv, os, re
from collections import defaultdict
from _paths import derived_path, raw_path

YEAR_RANGE = range(2010, 2027)  # 2010-2026


def parse_parents(field):
    if not field or str(field).strip() == '':
        return []
    parts = str(field).split(';')
    results = []
    for p in parts:
        p = p.strip()
        match = re.match(r'^(.+?)\s*\[(\d+\.?\d*)%\]$', p)
        if match:
            results.append((match.group(1).strip(), float(match.group(2))))
        elif p:
            results.append((p.strip(), None))
    return results


# 1. Load GEM->Compustat matches
parent_to_gvkeys = defaultdict(set)
with open(derived_path('mappings', 'gem_compustat_matches.csv'), 'r', encoding='utf-8') as f:
    for row in csv.DictReader(f):
        parent_to_gvkeys[row['gem_parent']].add(row['gvkey'])

print(f'Matched parents: {len(parent_to_gvkeys)}')

# 2. Read all plant-level data from GEM trackers
# For each plant: fuel type, capacity, ownership share, start year, retired year
FOSSIL_FUELS = {'coal', 'gas', 'oil'}

trackers = [
    ('Global-Coal-Plant-Tracker-January-2026.xlsx', 'Units', 'Parent', 'coal'),
    ('Global-Oil-and-Gas-Plant-Tracker-GOGPT-January-2026.xlsx', 'Gas & Oil Units', 'Parent(s)', 'gas'),
    ('Global-Solar-Power-Tracker-February-2026.xlsx', 'Utility-Scale (1 MW+)', 'Owner', 'solar'),
    ('Global-Wind-Power-Tracker-February-2026.xlsx', 'Data', 'Owner', 'wind'),
]

# gvkey -> list of (fuel, capacity_mw, start_year, retired_year_or_None)
gvkey_plants = defaultdict(list)

for fname, sheet, parent_col, fuel_type in trackers:
    fpath = raw_path('gem', fname)
    print(f'Reading {fname}...')
    wb = openpyxl.load_workbook(fpath, read_only=True)
    ws = wb[sheet]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers)}

    cap_col = 'Capacity (MW)'
    start_col = 'Start year'
    # Retired year column name varies
    ret_col = None
    for candidate in ['Retired year', 'Retired Year', 'retired year']:
        if candidate in col:
            ret_col = candidate
            break

    n_matched = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Include all statuses with a start year (operating, retired, mothballed)
        # We reconstruct the fleet at each point in time
        try:
            cap = float(row[col[cap_col]])
        except (ValueError, TypeError):
            continue

        # Start year
        try:
            start_yr = int(float(row[col[start_col]])) if row[col[start_col]] else None
        except (ValueError, TypeError):
            start_yr = None

        # Retired year
        ret_yr = None
        if ret_col and ret_col in col:
            try:
                ret_yr = int(float(row[col[ret_col]])) if row[col[ret_col]] else None
            except (ValueError, TypeError):
                ret_yr = None

        # Parse parents
        parsed = parse_parents(row[col[parent_col]])
        for name, pct in parsed:
            if name in parent_to_gvkeys:
                share = (pct / 100.0) if pct else 1.0 / len(parsed) if len(parsed) > 1 else 1.0
                for gvkey in parent_to_gvkeys[name]:
                    gvkey_plants[gvkey].append({
                        'fuel': fuel_type,
                        'mw': cap * share,
                        'start_year': start_yr,
                        'retired_year': ret_yr,
                    })
                    n_matched += 1

    wb.close()
    print(f'  Matched plant records: {n_matched}')

print(f'\nFirms with plant data: {len(gvkey_plants)}')

# 3. For each firm-year, compute alpha
print(f'\nComputing time-varying alpha for {len(YEAR_RANGE)} years...')
panel = []

for gvkey, plants in gvkey_plants.items():
    for year in YEAR_RANGE:
        coal_mw = 0.0
        gas_mw = 0.0
        fossil_mw = 0.0
        clean_mw = 0.0
        n_plants = 0

        for p in plants:
            # Is this plant operating in this year?
            # Operating if: start_year <= year AND (retired_year is None OR retired_year > year)
            start = p['start_year']
            ret = p['retired_year']

            if start is not None and start > year:
                continue  # not yet built
            if ret is not None and ret <= year:
                continue  # already retired

            # If start_year is None, assume operating (GEM lists it as current)
            n_plants += 1
            if p['fuel'] in FOSSIL_FUELS:
                fossil_mw += p['mw']
                if p['fuel'] == 'coal':
                    coal_mw += p['mw']
                elif p['fuel'] == 'gas':
                    gas_mw += p['mw']
            else:
                clean_mw += p['mw']

        total_mw = fossil_mw + clean_mw
        if total_mw > 0:
            alpha = fossil_mw / total_mw
            coal_share = coal_mw / total_mw
            panel.append({
                'gvkey': gvkey,
                'year': year,
                'coal_mw': f'{coal_mw:.1f}',
                'gas_mw': f'{gas_mw:.1f}',
                'fossil_mw': f'{fossil_mw:.1f}',
                'clean_mw': f'{clean_mw:.1f}',
                'total_mw': f'{total_mw:.1f}',
                'alpha': f'{alpha:.4f}',
                'coal_share': f'{coal_share:.4f}',
                'n_plants': n_plants,
            })

print(f'Panel rows: {len(panel)}')
print(f'Firms: {len(set(r["gvkey"] for r in panel))}')
print(f'Years: {len(set(r["year"] for r in panel))}')

# 4. Check for within-firm variation
firms_with_variation = 0
for gvkey in set(r['gvkey'] for r in panel):
    alphas = set(r['alpha'] for r in panel if r['gvkey'] == gvkey)
    if len(alphas) > 1:
        firms_with_variation += 1

total_firms = len(set(r['gvkey'] for r in panel))
print(f'\nFirms with alpha variation over time: {firms_with_variation} / {total_firms} ({100*firms_with_variation/total_firms:.1f}%)')

# Sample some firms with variation
print('\nSample firms with time-varying alpha:')
shown = 0
for gvkey in sorted(set(r['gvkey'] for r in panel)):
    rows = [(r['year'], float(r['alpha'])) for r in panel if r['gvkey'] == gvkey]
    alphas = [a for _, a in rows]
    if max(alphas) - min(alphas) > 0.05:  # meaningful variation
        first = rows[0]
        last = rows[-1]
        print(f'  {gvkey}: alpha {first[0]}={first[1]:.3f} -> {last[0]}={last[1]:.3f} (delta={last[1]-first[1]:+.3f})')
        shown += 1
        if shown >= 10:
            break

# 5. Save
outpath = derived_path('fundamentals', 'firm_alpha_panel.csv')
with open(outpath, 'w', newline='', encoding='utf-8') as f:
    w = csv.DictWriter(
        f,
        fieldnames=[
            'gvkey', 'year', 'coal_mw', 'gas_mw', 'fossil_mw', 'clean_mw',
            'total_mw', 'alpha', 'coal_share', 'n_plants'
        ],
    )
    w.writeheader()
    w.writerows(panel)

print(f'\nSaved {outpath}')
