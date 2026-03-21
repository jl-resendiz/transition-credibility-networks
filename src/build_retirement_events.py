"""Identify coal plant retirement events for the spatial event study.

Extracts retired coal units with year >= 2015, links to matched Compustat firms,
and identifies 'first-mover' events (first retirement in a geographic cluster).
"""
import openpyxl, csv, os, re, math, datetime
from collections import defaultdict

from _paths import raw_path, derived_path

def parse_parents(field):
    if not field or str(field).strip() == '':
        return []
    parts = str(field).split(';')
    results = []
    for p in parts:
        p = p.strip()
        match = re.match(r'^(.+?)\s*\[(\d+\.?\d*)%\]$', p)
        if match:
            results.append((match.group(1).strip(), float(match.group(2))))
        elif p:
            results.append((p.strip(), None))
    return results

def haversine(lat1, lon1, lat2, lon2):
    R = 6371
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dlat/2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon/2)**2
    return R * 2 * math.asin(math.sqrt(a))

# Load GEM->Compustat matches
parent_to_gvkeys = defaultdict(set)
with open(derived_path('mappings', 'gem_compustat_matches.csv'), 'r', encoding='utf-8') as f:
    for row in csv.DictReader(f):
        parent_to_gvkeys[row['gem_parent']].add(row['gvkey'])

# Read coal retirements
fpath = raw_path('gem', 'Global-Coal-Plant-Tracker-January-2026.xlsx')
wb = openpyxl.load_workbook(fpath, read_only=True)
ws = wb['Units']
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

col = {h: i for i, h in enumerate(headers)}

# Optional date columns (if GEM provides exact dates)
date_cols = [
    'Retired date',
    'Retirement date',
    'Retired Date',
    'Announcement date',
    'Announced retirement date',
]

retirements = []
for row in ws.iter_rows(min_row=2, values_only=True):
    status = str(row[col['Status']]) if row[col['Status']] else ''
    if status != 'retired':
        continue

    ret_year = row[col['Retired year']]
    if ret_year is None:
        continue
    try:
        ret_year = int(float(ret_year))
    except (ValueError, TypeError):
        continue
    if ret_year < 2015:
        continue

    try:
        lat = float(row[col['Latitude']])
        lon = float(row[col['Longitude']])
        cap = float(row[col['Capacity (MW)']])
    except (ValueError, TypeError):
        continue

    parent_raw = row[col['Parent']]
    parsed = parse_parents(parent_raw)

    # Check if any parent maps to Compustat
    matched_gvkeys = set()
    for name, pct in parsed:
        if name in parent_to_gvkeys:
            matched_gvkeys.update(parent_to_gvkeys[name])

    plant_name = row[col['Plant name']] or ''
    country = row[col['Country/Area']] or ''
    gem_id = row[col.get('GEM unit/phase ID', col.get('GEM phase ID', 0))] or ''

    # Exact event date if available
    event_date = ''
    for dc in date_cols:
        if dc in col:
            val = row[col[dc]]
            if val:
                if isinstance(val, (datetime.date, datetime.datetime)):
                    event_date = val.strftime('%Y-%m-%d')
                elif isinstance(val, (int, float)):
                    # Excel serial date (days since 1899-12-30)
                    base = datetime.date(1899, 12, 30)
                    try:
                        event_date = (base + datetime.timedelta(days=int(val))).strftime('%Y-%m-%d')
                    except Exception:
                        event_date = str(val)
                else:
                    event_date = str(val)
                    if len(event_date) >= 10:
                        event_date = event_date[:10]
                break

    retirements.append({
        'gem_id': str(gem_id),
        'plant_name': plant_name,
        'country': country,
        'parent_raw': str(parent_raw) if parent_raw else '',
        'ret_year': ret_year,
        'event_date': event_date,
        'capacity_mw': cap,
        'lat': lat,
        'lon': lon,
        'matched_gvkeys': ';'.join(sorted(matched_gvkeys)) if matched_gvkeys else '',
        'is_matched': len(matched_gvkeys) > 0,
    })

wb.close()

print(f'Coal retirements 2015+: {len(retirements)}')
print(f'Matched to Compustat: {sum(1 for r in retirements if r["is_matched"])}')

# Year distribution
from collections import Counter
year_dist = Counter(r['ret_year'] for r in retirements)
print(f'\nRetirements by year:')
for y in sorted(year_dist):
    matched_y = sum(1 for r in retirements if r['ret_year'] == y and r['is_matched'])
    print(f'  {y}: {year_dist[y]} total, {matched_y} matched')

# Country distribution
country_dist = Counter(r['country'] for r in retirements)
print(f'\nTop 10 countries:')
for c, n in country_dist.most_common(10):
    print(f'  {c}: {n}')

# Identify first-mover events: first retirement within 500km radius
RADIUS_KM = 500
retirements.sort(key=lambda r: r['ret_year'])

# Track which geographic cells have seen retirements
first_mover_events = []
retired_locations = []  # (lat, lon, year)

for r in retirements:
    lat, lon = r['lat'], r['lon']
    year = r['ret_year']

    # Check if any prior retirement within radius
    is_first = True
    for plat, plon, pyear in retired_locations:
        if pyear < year:
            d = haversine(lat, lon, plat, plon)
            if d <= RADIUS_KM:
                is_first = False
                break

    r['is_first_mover'] = is_first
    if is_first:
        first_mover_events.append(r)
    retired_locations.append((lat, lon, year))

print(f'\nFirst-mover events (first retirement in {RADIUS_KM}km radius): {len(first_mover_events)}')
print(f'First-movers matched to Compustat: {sum(1 for r in first_mover_events if r["is_matched"])}')

# Save all retirements
outpath = derived_path('events', 'coal_retirement_events.csv')
with open(outpath, 'w', newline='', encoding='utf-8') as f:
    fieldnames = ['gem_id', 'plant_name', 'country', 'parent_raw', 'ret_year', 'event_date',
                  'capacity_mw', 'lat', 'lon', 'matched_gvkeys', 'is_matched', 'is_first_mover']
    w = csv.DictWriter(f, fieldnames=fieldnames)
    w.writeheader()
    for r in retirements:
        w.writerow(r)

print(f'\nSaved {outpath} ({len(retirements)} events)')
