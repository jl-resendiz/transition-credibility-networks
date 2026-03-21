"""Parse all 4 GEM trackers, compute alpha (fossil intensity) per parent entity."""
import openpyxl, re, csv, os
from collections import defaultdict
from _paths import raw_path, derived_path

def parse_parents(field):
    """Parse 'AES Corp [100.0%]; Enel SpA [50.0%]' into [(name, pct), ...]"""
    if not field or str(field).strip() == '':
        return []
    parts = str(field).split(';')
    results = []
    for p in parts:
        p = p.strip()
        match = re.match(r'^(.+?)\s*\[(\d+\.?\d*)%\]$', p)
        if match:
            results.append((match.group(1).strip(), float(match.group(2))))
        elif p:
            results.append((p.strip(), None))
    return results

all_parents = defaultdict(lambda: {'coal_mw': 0, 'gas_mw': 0, 'solar_mw': 0, 'wind_mw': 0, 'units': 0})

trackers = [
    ('Global-Coal-Plant-Tracker-January-2026.xlsx', 'Units', 'Parent', 'Capacity (MW)', 'coal_mw'),
    ('Global-Oil-and-Gas-Plant-Tracker-GOGPT-January-2026.xlsx', 'Gas & Oil Units', 'Parent(s)', 'Capacity (MW)', 'gas_mw'),
    ('Global-Solar-Power-Tracker-February-2026.xlsx', 'Utility-Scale (1 MW+)', 'Owner', 'Capacity (MW)', 'solar_mw'),
    ('Global-Wind-Power-Tracker-February-2026.xlsx', 'Data', 'Owner', 'Capacity (MW)', 'wind_mw'),
]

for fname, sheet, parent_col, cap_col, mw_key in trackers:
    fpath = raw_path('gem', fname)
    print(f'Processing {fname}...')
    wb = openpyxl.load_workbook(fpath, read_only=True)
    ws = wb[sheet]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    parent_idx = headers.index(parent_col)
    cap_idx = headers.index(cap_col)
    status_idx = headers.index('Status')

    for row in ws.iter_rows(min_row=2, values_only=True):
        status = str(row[status_idx]) if row[status_idx] else ''
        if status != 'operating':
            continue
        cap = row[cap_idx]
        if cap is None:
            continue
        try:
            cap = float(cap)
        except (ValueError, TypeError):
            continue
        parsed = parse_parents(row[parent_idx])
        if not parsed:
            continue
        for name, pct in parsed:
            share = (pct / 100.0) if pct else 1.0 / len(parsed) if len(parsed) > 1 else 1.0
            all_parents[name][mw_key] += cap * share
            all_parents[name]['units'] += 1
    wb.close()

# Compute totals
for p in all_parents:
    all_parents[p]['fossil_mw'] = all_parents[p]['coal_mw'] + all_parents[p]['gas_mw']
    all_parents[p]['clean_mw'] = all_parents[p]['solar_mw'] + all_parents[p]['wind_mw']
    all_parents[p]['total_mw'] = all_parents[p]['fossil_mw'] + all_parents[p]['clean_mw']

sorted_parents = sorted(all_parents.items(), key=lambda x: x[1]['total_mw'], reverse=True)

print(f'\nTotal unique parent/owner entities: {len(all_parents)}')
print(f'\nTop 30 by total MW:')
fmt = '{:<45} {:>8} {:>8} {:>8} {:>8} {:>8} {:>6}'
print(fmt.format('Name', 'Coal', 'Gas', 'Solar', 'Wind', 'Total', 'Alpha'))
for name, d in sorted_parents[:30]:
    alpha = d['fossil_mw'] / d['total_mw'] if d['total_mw'] > 0 else 0
    print(fmt.format(name[:45], int(d['coal_mw']), int(d['gas_mw']), int(d['solar_mw']), int(d['wind_mw']), int(d['total_mw']), f'{alpha:.2f}'))

# Save full parent list
outpath = derived_path('mappings', 'gem_parents_parsed.csv')
with open(outpath, 'w', newline='', encoding='utf-8') as f:
    w = csv.writer(f)
    w.writerow(['parent_name', 'coal_mw', 'gas_mw', 'solar_mw', 'wind_mw', 'fossil_mw', 'clean_mw', 'total_mw', 'alpha', 'units'])
    for name, d in sorted_parents:
        alpha = d['fossil_mw'] / d['total_mw'] if d['total_mw'] > 0 else 0
        w.writerow([name, f"{d['coal_mw']:.1f}", f"{d['gas_mw']:.1f}", f"{d['solar_mw']:.1f}", f"{d['wind_mw']:.1f}",
                     f"{d['fossil_mw']:.1f}", f"{d['clean_mw']:.1f}", f"{d['total_mw']:.1f}", f'{alpha:.4f}', d['units']])

print(f'\nSaved {outpath} ({len(all_parents)} entities)')
