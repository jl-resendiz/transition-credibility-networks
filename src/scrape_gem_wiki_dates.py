"""Scrape GEM Wiki pages to extract exact retirement dates for first-mover events.

For each matched first-mover coal retirement, fetches the GEM wiki page and
extracts the most precise retirement date available from the page text.

Output: gem_wiki_dates.csv with columns: gem_id, plant_name, wiki_url, extracted_date, date_source, raw_snippet
"""
import csv, os, re, time
import urllib.request
import urllib.error
import openpyxl
from collections import defaultdict
from html.parser import HTMLParser

from _paths import raw_path, derived_path


# Simple HTML-to-text parser
class HTMLToText(HTMLParser):
    def __init__(self):
        super().__init__()
        self.text = []
        self.skip = False

    def handle_starttag(self, tag, attrs):
        if tag in ('script', 'style', 'noscript'):
            self.skip = True

    def handle_endtag(self, tag):
        if tag in ('script', 'style', 'noscript'):
            self.skip = False
        if tag in ('p', 'br', 'div', 'li', 'h1', 'h2', 'h3', 'h4', 'td', 'th', 'tr'):
            self.text.append('\n')

    def handle_data(self, data):
        if not self.skip:
            self.text.append(data)

    def get_text(self):
        return ''.join(self.text)


def html_to_text(html):
    parser = HTMLToText()
    parser.feed(html)
    return parser.get_text()


# 1. Load GEM tracker to map gem_id -> wiki_url
print('Loading GEM coal tracker for Wiki URLs...')
fpath = raw_path('gem', 'Global-Coal-Plant-Tracker-January-2026.xlsx')
wb = openpyxl.load_workbook(fpath, read_only=True)
ws = wb['Units']
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
col = {h: i for i, h in enumerate(headers)}

gem_wiki = {}  # gem_id -> wiki_url
gem_plant_names = {}  # gem_id -> plant_name
for row in ws.iter_rows(min_row=2, values_only=True):
    gid = str(row[col['GEM unit/phase ID']]) if row[col['GEM unit/phase ID']] else ''
    url = str(row[col['Wiki URL']]) if row[col['Wiki URL']] else ''
    pname = str(row[col['Plant name']]) if row[col['Plant name']] else ''
    if gid and url:
        gem_wiki[gid] = url
        gem_plant_names[gid] = pname
wb.close()
print(f'  GEM units with Wiki URLs: {len(gem_wiki)}')

# 2. Load matched first-mover events
events = []
with open(derived_path('events', 'coal_retirement_events.csv'), 'r', encoding='utf-8') as f:
    for row in csv.DictReader(f):
        if row['is_first_mover'] == 'True' and row['is_matched'] == 'True':
            events.append(row)

print(f'  Matched first-mover events: {len(events)}')

# Deduplicate by wiki URL (multiple units at same plant share URL)
url_to_events = defaultdict(list)
for e in events:
    gid = e['gem_id']
    url = gem_wiki.get(gid, '')
    if url:
        url_to_events[url].append(e)

print(f'  Unique Wiki URLs to scrape: {len(url_to_events)}')

# 3. Date extraction patterns
MONTHS = {
    'january': '01', 'february': '02', 'march': '03', 'april': '04',
    'may': '05', 'june': '06', 'july': '07', 'august': '08',
    'september': '09', 'october': '10', 'november': '11', 'december': '12',
    'jan': '01', 'feb': '02', 'mar': '03', 'apr': '04',
    'jun': '06', 'jul': '07', 'aug': '08', 'sep': '09',
    'oct': '10', 'nov': '11', 'dec': '12',
}

# Keywords that signal retirement-related context
RETIRE_KEYWORDS = [
    'retire', 'retired', 'retirement', 'retir',
    'close', 'closed', 'closure', 'closing',
    'shut down', 'shutdown', 'shut-down',
    'decommission', 'decommissioned',
    'cease', 'ceased', 'ceasing',
    'demolish', 'demolished', 'demolition',
    'taken offline', 'went offline', 'taken off-line',
    'last day', 'final day', 'last unit',
    'phased out', 'phase out', 'phase-out',
    'permanently', 'no longer',
]

# Pattern: "DD Month YYYY" or "Month DD, YYYY" near retirement keywords
DATE_PATTERNS = [
    # "31 March 2021", "1 January 2020"
    re.compile(r'(\d{1,2})\s+(january|february|march|april|may|june|july|august|september|october|november|december)\s+(\d{4})', re.I),
    # "March 31, 2021"
    re.compile(r'(january|february|march|april|may|june|july|august|september|october|november|december)\s+(\d{1,2}),?\s+(\d{4})', re.I),
    # "in March 2021", "by December 2020"
    re.compile(r'(?:in|by|on|during|since|from|until|before|after)\s+(january|february|march|april|may|june|july|august|september|october|november|december)\s+(\d{4})', re.I),
    # ISO-like: "2021-03-31" or "2021/03/31"
    re.compile(r'(\d{4})[-/](\d{2})[-/](\d{2})'),
]


def extract_retirement_date(text, ret_year):
    """Extract the most precise retirement date from page text.

    Strategy:
    1. Find paragraphs/sentences containing retirement keywords
    2. Extract dates from those contexts
    3. Prefer dates matching the retirement year
    4. Prefer more precise dates (day+month > month only)
    """
    text_lower = text.lower()
    ret_year_str = str(ret_year)

    # Find sentences containing retirement keywords
    sentences = re.split(r'[.\n]', text)
    retire_sentences = []
    for sent in sentences:
        sent_lower = sent.lower()
        for kw in RETIRE_KEYWORDS:
            if kw in sent_lower:
                retire_sentences.append(sent)
                break

    if not retire_sentences:
        # Fallback: search entire text
        retire_sentences = sentences

    best_date = None
    best_precision = 0  # 1=year+month, 2=year+month+day
    best_snippet = ''

    for sent in retire_sentences:
        sent_stripped = sent.strip()
        if len(sent_stripped) < 5:
            continue

        # Pattern 1: "DD Month YYYY"
        for m in DATE_PATTERNS[0].finditer(sent):
            day, month_str, year = m.group(1), m.group(2).lower(), m.group(3)
            if year == ret_year_str or abs(int(year) - ret_year) <= 1:
                mm = MONTHS.get(month_str, '')
                if mm:
                    date = f'{year}-{mm}-{int(day):02d}'
                    if best_precision < 2 or year == ret_year_str:
                        best_date = date
                        best_precision = 2
                        best_snippet = sent_stripped[:200]

        # Pattern 2: "Month DD, YYYY"
        for m in DATE_PATTERNS[1].finditer(sent):
            month_str, day, year = m.group(1).lower(), m.group(2), m.group(3)
            if year == ret_year_str or abs(int(year) - ret_year) <= 1:
                mm = MONTHS.get(month_str, '')
                if mm:
                    date = f'{year}-{mm}-{int(day):02d}'
                    if best_precision < 2 or year == ret_year_str:
                        best_date = date
                        best_precision = 2
                        best_snippet = sent_stripped[:200]

        # Pattern 3: "in Month YYYY" (month precision only)
        if best_precision < 2:
            for m in DATE_PATTERNS[2].finditer(sent):
                month_str, year = m.group(1).lower(), m.group(2)
                if year == ret_year_str or abs(int(year) - ret_year) <= 1:
                    mm = MONTHS.get(month_str, '')
                    if mm:
                        date = f'{year}-{mm}-15'  # mid-month as proxy
                        if best_precision < 1 or year == ret_year_str:
                            best_date = date
                            best_precision = 1
                            best_snippet = sent_stripped[:200]

        # Pattern 4: ISO date
        for m in DATE_PATTERNS[3].finditer(sent):
            year, mm, dd = m.group(1), m.group(2), m.group(3)
            if year == ret_year_str or abs(int(year) - ret_year) <= 1:
                date = f'{year}-{mm}-{dd}'
                if best_precision < 2 or year == ret_year_str:
                    best_date = date
                    best_precision = 2
                    best_snippet = sent_stripped[:200]

    return best_date, best_precision, best_snippet


# 4. Scrape Wiki pages
print(f'\nScraping {len(url_to_events)} GEM Wiki pages...')
results = []
success = 0
failed = 0
dates_found = 0

urls = sorted(url_to_events.keys())
for idx, url in enumerate(urls):
    evts = url_to_events[url]
    ret_year = int(evts[0]['ret_year'])
    plant = evts[0]['plant_name']

    if (idx + 1) % 10 == 0 or idx == 0:
        print(f'  [{idx+1}/{len(urls)}] {plant}...')

    try:
        req = urllib.request.Request(url, headers={
            'User-Agent': 'Mozilla/5.0 (research; Oxford University academic project)',
            'Accept': 'text/html',
        })
        with urllib.request.urlopen(req, timeout=15) as resp:
            html = resp.read().decode('utf-8', errors='replace')
        text = html_to_text(html)
        date, precision, snippet = extract_retirement_date(text, ret_year)

        for e in evts:
            result = {
                'gem_id': e['gem_id'],
                'plant_name': e['plant_name'],
                'country': e['country'],
                'ret_year': e['ret_year'],
                'wiki_url': url,
                'extracted_date': date or '',
                'date_precision': 'day' if precision == 2 else ('month' if precision == 1 else ''),
                'raw_snippet': snippet,
            }
            results.append(result)

        if date:
            dates_found += 1
        success += 1

    except (urllib.error.URLError, urllib.error.HTTPError, OSError, TimeoutError) as exc:
        for e in evts:
            results.append({
                'gem_id': e['gem_id'],
                'plant_name': e['plant_name'],
                'country': e['country'],
                'ret_year': e['ret_year'],
                'wiki_url': url,
                'extracted_date': '',
                'date_precision': '',
                'raw_snippet': f'FETCH ERROR: {exc}',
            })
        failed += 1

    # Polite delay
    time.sleep(0.5)

print(f'\nScraping complete:')
print(f'  Pages fetched: {success}')
print(f'  Pages failed: {failed}')
print(f'  Dates extracted: {dates_found} / {len(urls)} unique URLs')
print(f'  Events covered: {len(results)}')

# Precision breakdown
day_count = sum(1 for r in results if r['date_precision'] == 'day')
month_count = sum(1 for r in results if r['date_precision'] == 'month')
none_count = sum(1 for r in results if r['date_precision'] == '')
print(f'  Day precision: {day_count}')
print(f'  Month precision: {month_count}')
print(f'  No date found: {none_count}')

# 5. Save results
outpath = derived_path('events', 'gem_wiki_dates.csv')
with open(outpath, 'w', newline='', encoding='utf-8') as f:
    w = csv.DictWriter(f, fieldnames=[
        'gem_id', 'plant_name', 'country', 'ret_year',
        'wiki_url', 'extracted_date', 'date_precision', 'raw_snippet',
    ])
    w.writeheader()
    w.writerows(results)

print(f'\nSaved {outpath}')

# 6. Update coal_retirement_events.csv with extracted dates
print('\nUpdating coal_retirement_events.csv with extracted dates...')
date_lookup = {r['gem_id']: r['extracted_date'] for r in results if r['extracted_date']}
print(f'  Date lookup entries: {len(date_lookup)}')

all_events = []
with open(derived_path('events', 'coal_retirement_events.csv'), 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    fieldnames = reader.fieldnames
    for row in reader:
        if row['gem_id'] in date_lookup and not row.get('event_date'):
            row['event_date'] = date_lookup[row['gem_id']]
        all_events.append(row)

updated = sum(1 for e in all_events if e.get('event_date'))
with open(derived_path('events', 'coal_retirement_events.csv'), 'w', newline='', encoding='utf-8') as f:
    w = csv.DictWriter(f, fieldnames=fieldnames)
    w.writeheader()
    w.writerows(all_events)

print(f'  Events with exact date: {updated}')
print(f'  Updated coal_retirement_events.csv')
